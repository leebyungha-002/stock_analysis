import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import io
import os
import sys
import re
import unicodedata
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

# =============================================================================
# 🛠️ [1. 사용자 기본 설정]
# =============================================================================

USER_TOP_ANALYSIS_CONFIG = [
    ('보통예금', 20),
    ('외상매출금', 10),
    ('미지급금', 10),
    ('지급수수료', 10),
    ('복리후생비', 10)
]

RELATED_PARTIES = [
    '(주)세중샤론손해보험', '대표이사', '(주)세중엔지니어링',
    '(재)우리옛돌문화재단', '주)세중인터내셔널', '(주)세중인터내셔널', '(주)세성항운',
    '천신일', '천세전', '천호전', '천미전'
]

SEARCH_KEYWORDS = [
    '상품권', '접대', '가수금', '선물', '회식', 
    '결산수정', '비자금', '가지급', '리베이트', '현금'
]

COL_DATE = '전표일자'
COL_JOURNAL_ID = '전표번호'
COL_ACCOUNT = '계정명'
COL_CLIENT = '거래처명'
COL_DESC = '적요'
COL_DEBIT = '차변'
COL_CREDIT = '대변'
COL_EMPLOYEE = '사원명'  # 사원명 컬럼

MASK_TARGET_COLS = ['적요', '관리항목4', '비고', '내용']
DEFAULT_BENFORD_TARGETS = [('복리후생비', '차변'), ('접대비', '차변'), ('여비교통비', '차변')]

GLOBAL_SAFE_MAP = {}
CLIENT_COUNTER = 1

# 결과 저장·데이터 로드 기준 경로 (이 스크립트가 있는 폴더, 실행 위치와 무관)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def _has_expected_columns(df):
    """차변/대변/계정/전표일자/거래처 등 분석에 필요한 컬럼이 하나라도 있는지 확인"""
    cols = [str(c).strip() for c in df.columns]
    keywords = ('차변', '대변', '계정', '전표일자', '거래처', '전표번호', '적요')
    return any(any(kw in c for kw in keywords) for c in cols)

def _read_excel_with_header_detection(path, dtype_map=None):
    """엑셀을 0, 1, 2행을 헤더 후보로 시도해 인식된 (df, 사용한_헤더행) 반환. 헤더행은 0/1/2."""
    dtype_map = dtype_map or {}
    for header_row in range(3):
        try:
            trial = pd.read_excel(path, engine='openpyxl', header=header_row, dtype=dtype_map)
            if trial.empty or len(trial.columns) < 2:
                continue
            if _has_expected_columns(trial):
                return trial, header_row
        except Exception:
            continue
    df = pd.read_excel(path, engine='openpyxl', dtype=dtype_map)
    return df, 0

# =============================================================================
# 🔧 [보조 함수]
# =============================================================================

def get_first_digit(number):
    try:
        s = str(abs(int(number)))
        return int(s[0]) if s[0] != '0' else 0
    except: return 0

def find_data_file():
    target_dir = os.getcwd()
    files = os.listdir(target_dir)
    data_files = []
    for f in files:
        if (f.endswith('.csv') or f.endswith('.xlsx')) and not f.startswith('~$') and not f.startswith('JET_') and not f.startswith('AI_') and not f.endswith('.png'):
            data_files.append(f)
    if not data_files: return None
    data_files.sort(key=lambda x: os.path.getmtime(os.path.join(target_dir, x)), reverse=True)
    return data_files[0]

def load_data():
    """
    data/current/ 와 data/previous/ 폴더를 스캔하여 csv/xlsx 파일을 읽고,
    '구분' 컬럼(당기/전기)을 붙여 합친 DataFrame을 반환.
    기준 경로: 이 스크립트(analysis.py)가 있는 폴더.
    """
    dtype_map = {COL_JOURNAL_ID: str}
    base_dir = SCRIPT_DIR
    current_dir = os.path.join(base_dir, 'data', 'current')
    previous_dir = os.path.join(base_dir, 'data', 'previous')
    all_dfs = []

    def read_file(path, filename):
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == '.csv':
                print(f"     📂 CSV 로드 중: {filename} ...", flush=True)
                encodings = ['utf-8', 'cp949', 'utf-16', 'euc-kr', 'latin-1']
                out = None
                for enc in encodings:
                    try:
                        out = pd.read_csv(path, encoding=enc, dtype=dtype_map)
                        break
                    except UnicodeDecodeError:
                        continue
                if out is None:
                    raise ValueError("지원하는 인코딩(utf-8, cp949, utf-16, euc-kr, latin-1)으로 CSV를 읽을 수 없습니다.")
                print(f"     ✅ CSV 로드 완료: {len(out)}행", flush=True)
                return out
            elif ext == '.xlsx':
                print(f"     📂 엑셀 로드 중: {filename} ...", flush=True)
                out, header_row = _read_excel_with_header_detection(path, dtype_map)
                if header_row > 0:
                    print(f"     ℹ️ 헤더를 {header_row + 1}번째 행으로 인식했습니다.", flush=True)
                print(f"     ✅ 엑셀 로드 완료: {len(out)}행", flush=True)
                return out
        except PermissionError:
            print(f"     ⚠️ 파일 읽기 실패: {filename}", flush=True)
            print(f"     💡 해당 파일이 Excel 등에서 열려 있습니다. 데이터 파일을 닫고 다시 실행하세요.", flush=True)
            return None
        except Exception as e:
            err_msg = str(e).lower()
            if 'permission' in err_msg or 'being used' in err_msg or '다른 프로세스' in err_msg:
                print(f"     ⚠️ 파일 읽기 실패: {filename}", flush=True)
                print(f"     💡 해당 파일이 Excel 등에서 열려 있습니다. 데이터 파일을 닫고 다시 실행하세요.", flush=True)
            else:
                print(f"     ⚠️ 파일 읽기 실패: {path} - {e}", flush=True)
            return None

    # 당기: data/current/
    if os.path.isdir(current_dir):
        for f in sorted(os.listdir(current_dir)):
            if f.endswith('.csv') or f.endswith('.xlsx'):
                if f.startswith('~$'): continue
                path = os.path.join(current_dir, f)
                df = read_file(path, f)
                if df is not None and not df.empty:
                    df = df.copy()
                    df['구분'] = '당기'
                    all_dfs.append(df)
                    print(f"   📂 [당기] 로드: {f} ({len(df)}행)")
    else:
        print(f"   ℹ️ 폴더 없음 (당기): {current_dir}")

    # 전기: data/previous/
    if os.path.isdir(previous_dir):
        for f in sorted(os.listdir(previous_dir)):
            if f.endswith('.csv') or f.endswith('.xlsx'):
                if f.startswith('~$'): continue
                path = os.path.join(previous_dir, f)
                df = read_file(path, f)
                if df is not None and not df.empty:
                    df = df.copy()
                    df['구분'] = '전기'
                    all_dfs.append(df)
                    print(f"   📂 [전기] 로드: {f} ({len(df)}행)")
    else:
        print(f"   ℹ️ 폴더 없음 (전기): {previous_dir}")

    if not all_dfs:
        return None
    return pd.concat(all_dfs, axis=0, ignore_index=True)

def get_safe_client_name(real_name):
    global CLIENT_COUNTER
    if pd.isna(real_name) or str(real_name).strip() == '': return "(미기재)"
    real_name = str(real_name).strip()
    if real_name in GLOBAL_SAFE_MAP: return GLOBAL_SAFE_MAP[real_name]
    else:
        alias = f"Client_{CLIENT_COUNTER:03d}"
        GLOBAL_SAFE_MAP[real_name] = alias
        CLIENT_COUNTER += 1
        return alias

def mask_sensitive_info(text):
    if not isinstance(text, str):
        text = str(text) if text is not None else ""
        if text == "": return text
    hyphen_pattern = r'(\d{2,})[-](\d{2,})[-](\d{2,})'
    text = re.sub(hyphen_pattern, r'\1-****-\3', text)
    def mask_long_num(match):
        full_num = match.group()
        if len(full_num) >= 10: return full_num[:4] + "****" + full_num[-4:]
        return full_num
    no_hyphen_pattern = r'\b\d{10,}\b'
    text = re.sub(no_hyphen_pattern, mask_long_num, text)
    return text

def draw_benford_chart(account_name, direction, actual_probs, benford_probs):
    """벤포드 차트를 메모리에 그려 Excel 삽입용 버퍼를 반환 (파일로 저장하지 않음)"""
    try:
        plt.rc('font', family='Malgun Gothic')
        plt.rcParams['axes.unicode_minus'] = False
        digits = range(1, 10)
        actual_values = [actual_probs.get(d, 0.0) * 100 for d in digits]
        benford_values = [benford_probs[d] * 100 for d in digits]
        plt.figure(figsize=(10, 6))
        plt.plot(digits, benford_values, color='red', marker='o', linestyle='--', label='벤포드 법칙')
        plt.bar(digits, actual_values, color='skyblue', alpha=0.7, label=f'실제 ({account_name})')
        plt.title(f'벤포드 분석: {account_name} ({direction})')
        plt.legend()
        plt.grid(axis='y', linestyle='--', alpha=0.5)
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        plt.close()
        buf.seek(0)
        return buf
    except Exception:
        return None

# =============================================================================
# 📊 [분석 로직 함수들]
# =============================================================================

def run_data_overview(df, writer):
    print("   ▶ [0. Overview] 상세 통계 작성 중...")
    min_date = df[COL_DATE].min()
    max_date = df[COL_DATE].max()
    summary = {
        '구분': ['총 행수', '총 차변', '총 대변', '계정 수', '시작일', '종료일'],
        '값': [len(df), df[COL_DEBIT].sum(), df[COL_CREDIT].sum(), df[COL_ACCOUNT].nunique(), min_date, max_date]
    }
    pd.DataFrame(summary).to_excel(writer, sheet_name='0_데이터개요', index=False)
    debit_stats = df[df[COL_DEBIT] != 0].groupby(COL_ACCOUNT)[COL_DEBIT].agg(['count', 'sum', 'mean', 'std'])
    debit_stats.columns = ['전표건수(차)', '차변합계', '평균금액(차)', '표준편차(차)']
    credit_stats = df[df[COL_CREDIT] != 0].groupby(COL_ACCOUNT)[COL_CREDIT].agg(['count', 'sum', 'mean', 'std'])
    credit_stats.columns = ['전표건수(대)', '대변합계', '평균금액(대)', '표준편차(대)']
    full_stats = pd.concat([debit_stats, credit_stats], axis=1).fillna(0).sort_values(by='차변합계', ascending=False)
    pd.DataFrame({'제목': ['[계정과목별 상세 통계표]']}).to_excel(writer, sheet_name='0_데이터개요', startrow=8, index=False, header=False)
    full_stats.to_excel(writer, sheet_name='0_데이터개요', startrow=9)

# 벤포드 분석 최소 건수 (이 건수 미만이면 분석 생략, 시트에는 '데이터 부족' 안내만 기록)
BENFORD_MIN_ROWS = 5

def _normalize_account_for_match(s):
    """계정명 비교용 정규화: 괄호·공백 제거, 소문자 통일(S&C↔s&c). 매출액(상품)↔매출액상품 동일 매칭."""
    s = str(s).strip()
    # 전각·CJK 괄호 등 → 반각 ( ) 로 통일
    for full, half in [
        ('\uff08', '('), ('\uff09', ')'),   # （） 전각
        ('\u2985', '('), ('\u2986', ')'),   # ⓵⓶
        ('\ufe59', '('), ('\ufe5a', ')'),   # ﹙﹚ small
        ('\u27ee', '('), ('\u27ed', ')'),   # ⟩⟨
        ('\u3008', '('), ('\u3009', ')'),   # 〈〉 angle
        ('\u3010', '('), ('\u3011', ')'),   # 【】 black lenticular
    ]:
        s = s.replace(full, half)
    s = s.replace('\u00a0', ' ')  # non-breaking space → 일반 공백
    s = s.replace('\uff06', '&')  # 전각 ＆ → 반각 & (용역매출액(S&C) 등)
    s = re.sub(r'\s+', '', s)    # 모든 공백 제거
    s = re.sub(r'[()（）\uff08\uff09]+', '', s)  # 괄호 제거 → 매출액(상품) → 매출액상품
    s = s.replace('권', '')      # 기존 완화 (국제항공권 등)
    s = s.lower()  # S&C ↔ s&c, 용역/상품 구분 유지
    return s

def _account_match_flexible(acct_series, acct_str):
    """
    계정명 매칭: ( ) 가 붙은 계정명(미수금(기타) 등) 인식.
    - 괄호 전각/반각/공백 차이 무시 후 contains
    - 0건이면 '괄호 앞부분'만으로도 포함 검사 (미수금(기타) → 미수금 포함 시 매칭)
    """
    acct_str = str(acct_str).strip()
    if not acct_str:
        return pd.Series(False, index=acct_series.index)
    norm_series = acct_series.fillna('').astype(str).apply(_normalize_account_for_match)
    norm_user = _normalize_account_for_match(acct_str)

    # 1) 데이터 계정명이 사용자 입력으로 시작 → 용역매출액은 상품매출액과 섞이지 않음
    mask_start = norm_series.str.startswith(norm_user)
    if mask_start.any():
        return mask_start

    # 2) 포함 여부 (contains)
    mask_contains = norm_series.str.contains(re.escape(norm_user), na=False, regex=True, case=False)
    if mask_contains.any():
        return mask_contains

    # 3) 괄호 앞부분 등 보조 매칭
    def _row_match(val):
        v = str(val) if pd.notna(val) else ''
        if not v:
            return False
        nv = _normalize_account_for_match(v)
        if norm_user in nv or nv in norm_user:
            return True
        if len(norm_user) >= 2 and nv.startswith(norm_user):
            return True
        if len(norm_user) >= 2 and norm_user.startswith(nv):
            return True
        return False
    return acct_series.fillna('').apply(_row_match)

BENFORD_EXPLANATION = [
    '[벤포드 분석의 의미]',
    '벤포드 법칙: 자연발생한 숫자 데이터에서 맨 앞자리(선행자릿수) 1~9의 출현 비율이 일정한 패턴을 따른다는 법칙입니다.',
    '1이 약 30.1%, 2가 약 17.6%, 3이 약 12.5%… 9가 약 4.6%로 점점 감소하는 것이 정상입니다.',
    '회계·재무 데이터에서 이 패턴에서 크게 벗어나면, 위조·조작·이상거래 가능성을 시사할 수 있어 회계감사에서 활용됩니다.',
    '실제비율과 이론비율의 차이가 클수록 벤포드 법칙에서 이탈한 것으로 해석할 수 있습니다.',
]

def run_monthly_full_account(df, writer):
    """계정별·월별 차변/대변 합계를 시트에 기록"""
    df = df.copy()
    df['Month'] = pd.to_datetime(df[COL_DATE], errors='coerce').dt.month
    monthly_grp = df.groupby([COL_ACCOUNT, 'Month'])[[COL_DEBIT, COL_CREDIT]].sum().reset_index()
    monthly_grp.to_excel(writer, sheet_name='2_월별_전계정_분석', index=False)

def run_jet_analysis(df, writer, benford_targets):
    print(f"   ▶ [1. JET] 벤포드 분석 수행")
    run_monthly_full_account(df, writer)
    benford_images = []
    benford_probs = {1:0.301, 2:0.176, 3:0.125, 4:0.097, 5:0.079, 6:0.067, 7:0.058, 8:0.051, 9:0.046}
    pd.DataFrame({'내용': BENFORD_EXPLANATION}).to_excel(writer, sheet_name='6_벤포드분석', startrow=0, index=False, header=False)
    row_cursor = len(BENFORD_EXPLANATION) + 2
    for acct, direction in benford_targets:
        target_col = COL_DEBIT if direction == '차변' else COL_CREDIT
        mask = _account_match_flexible(df[COL_ACCOUNT], acct)
        subset = df[mask & (df[target_col] > 0)].copy()
        n = len(subset)
        if n < BENFORD_MIN_ROWS:
            # 데이터 부족 시 시트에 안내 행 기록 + 콘솔 메시지
            skip_df = pd.DataFrame([{'계정': acct, '방향': direction, '비고': f'데이터 부족 ({n}건, {BENFORD_MIN_ROWS}건 이상 필요)'}])
            skip_df.to_excel(writer, sheet_name='6_벤포드분석', startrow=row_cursor, index=False)
            row_cursor += 4
            print(f"     ⚠️ [{acct} ({direction})] 데이터 {n}건 → {BENFORD_MIN_ROWS}건 미만으로 분석 생략")
            continue
        subset['Digit'] = subset[target_col].apply(get_first_digit)
        digit_subset = subset[subset['Digit'] >= 1]['Digit']
        counts = digit_subset.value_counts(normalize=True).sort_index()
        counts_raw = digit_subset.value_counts().sort_index()
        img = draw_benford_chart(acct, direction, counts, benford_probs)
        if img: benford_images.append(img)
        res = []
        for d in range(1, 10):
            actual = counts.get(d, 0.0)
            theory = benford_probs[d]
            diff = round(actual - theory, 3)
            res.append({
                '계정': acct, '방향': direction, '숫자': d,
                '발생건수': int(counts_raw.get(d, 0)),
                '실제비율': round(actual, 3),
                '이론비율': theory,
                '차이(실제-이론)': diff
            })
        if res:
            pd.DataFrame(res).to_excel(writer, sheet_name='6_벤포드분석', startrow=row_cursor, index=False)
            row_cursor += 15
    return benford_images

def run_keyword_search(df, writer, keywords):
    if not keywords: return
    print(f"   ▶ [2. 키워드] 적요 검색 중...")
    if COL_DESC not in df.columns: return
    safe_keywords = [re.escape(k) for k in keywords]
    pattern = '|'.join(safe_keywords)
    mask = df[COL_DESC].str.contains(pattern, na=False, regex=True)
    keyword_df = df[mask].copy()
    sheet_name = '7_키워드검색'
    if not keyword_df.empty:
        keyword_df['AbsAmt'] = keyword_df[[COL_DEBIT, COL_CREDIT]].abs().max(axis=1)
        keyword_df = keyword_df.sort_values(by='AbsAmt', ascending=False).drop(columns=['AbsAmt'])
        info = [f"검색어: {', '.join(keywords)}", f"발견 건수: {len(keyword_df)}건"]
        pd.DataFrame({'내용': info}).to_excel(writer, sheet_name=sheet_name, startrow=0, index=False, header=False)
        keyword_df.to_excel(writer, sheet_name=sheet_name, startrow=3, index=False)
    else:
        pd.DataFrame({'결과': [f"검색어 '{', '.join(keywords)}' 포함 전표 없음"]}).to_excel(writer, sheet_name=sheet_name, index=False)

def run_round_number_analysis(df, writer):
    print("   ▶ [3. 라운드] 라운드 넘버 분석")
    target = df[df[COL_DEBIT] > 0]
    round_1m = target[target[COL_DEBIT] % 1000000 == 0]
    if not round_1m.empty:
        round_1m.nlargest(1000, COL_DEBIT).to_excel(writer, sheet_name='8_라운드넘버', index=False)

def _related_party_pattern(party):
    """(주)로 시작하면 (주) 없어도 매칭: (주)세중샤론 → 세중샤론 포함"""
    p = str(party).strip()
    if p.startswith('(주)'):
        return r'(?:\(주\))?' + re.escape(p[3:])
    return re.escape(p)

def run_related_party_analysis(df, writer, parties=None):
    """parties: 분석 대상 특수관계자 리스트 (None이면 RELATED_PARTIES 사용)"""
    # 이전 버전에서 생성된 분리 시트 제거 (현재는 하나의 시트로 통합)
    if hasattr(writer, 'book') and writer.book is not None and '5_특수관계자_계정별요약' in writer.book.sheetnames:
        del writer.book['5_특수관계자_계정별요약']
    party_list = parties if parties is not None else RELATED_PARTIES
    print(f"   ▶ [4. 특수관계자] {len(party_list)}개 지정 거래처 분석 중...")
    if COL_CLIENT not in df.columns: return
    safe_parties = [_related_party_pattern(p) for p in party_list]
    pattern = '|'.join(safe_parties)
    related_df = df[df[COL_CLIENT].str.contains(pattern, na=False, regex=True)].copy()
    sheet_name = '5_특수관계자거래'
    if not related_df.empty:
        pivot_debit = related_df.pivot_table(index=COL_ACCOUNT, columns=COL_CLIENT, values=COL_DEBIT, aggfunc='sum', fill_value=0).reset_index()
        pivot_credit = related_df.pivot_table(index=COL_ACCOUNT, columns=COL_CLIENT, values=COL_CREDIT, aggfunc='sum', fill_value=0).reset_index()
        rel_summary = related_df.groupby([COL_CLIENT, COL_ACCOUNT])[[COL_DEBIT, COL_CREDIT]].agg(['sum', 'count']).reset_index()
        rel_summary.columns = ['거래처명', '계정명', '차변합계', '차변건수', '대변합계', '대변건수']

        # openpyxl로 직접 기록 (pandas 다중 to_excel 시 미표시 이슈 방지)
        wb = writer.book
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            wb.remove(ws)
        ws = wb.create_sheet(sheet_name, 0)
        row = 1
        ws.cell(row=row, column=1, value="[계정별·특수관계자별 요약] 차변합계")
        row += 1
        for r_idx, r_row in enumerate(dataframe_to_rows(pivot_debit, index=False, header=True), row):
            for c_idx, val in enumerate(r_row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        row += len(pivot_debit) + 4
        ws.cell(row=row, column=1, value="[계정별·특수관계자별 요약] 대변합계")
        row += 1
        for r_idx, r_row in enumerate(dataframe_to_rows(pivot_credit, index=False, header=True), row):
            for c_idx, val in enumerate(r_row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        row += len(pivot_credit) + 4
        ws.cell(row=row, column=1, value=f"특수관계자 거래 내역 ({len(related_df)}건)")
        row += 2
        for r_idx, r_row in enumerate(dataframe_to_rows(rel_summary, index=False, header=True), row):
            for c_idx, val in enumerate(r_row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        row += len(rel_summary) + 4
        ws.cell(row=row, column=1, value="상세 거래 내역 (전체)")
        row += 1
        for r_idx, r_row in enumerate(dataframe_to_rows(related_df, index=False, header=True), row):
            for c_idx, val in enumerate(r_row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    else:
        pd.DataFrame({'결과': ['지정된 특수관계자와의 거래 내역이 없습니다.']}).to_excel(writer, sheet_name=sheet_name, index=False)

def run_asset_liability_cross_check(df, writer, assets, liabilities):
    if not assets or not liabilities: return
    print("   ▶ [5. 교차분석] 자산 vs 부채")
    asset_mask = pd.Series([False] * len(df))
    for a in assets: asset_mask |= df[COL_ACCOUNT].str.contains(a, na=False, regex=False)
    asset_df = df[asset_mask & (df[COL_DEBIT] > 0)]
    liab_mask = pd.Series([False] * len(df))
    for l in liabilities: liab_mask |= df[COL_ACCOUNT].str.contains(l, na=False, regex=False)
    liab_df = df[liab_mask & (df[COL_CREDIT] > 0)]
    if asset_df.empty or liab_df.empty: return
    grp_asset = asset_df.groupby(COL_CLIENT).agg({COL_DEBIT: 'sum', COL_ACCOUNT: lambda x: ','.join(set(x))}).reset_index()
    grp_asset.columns = ['거래처명', '자산_금액', '자산_계정']
    grp_liab = liab_df.groupby(COL_CLIENT).agg({COL_CREDIT: 'sum', COL_ACCOUNT: lambda x: ','.join(set(x))}).reset_index()
    grp_liab.columns = ['거래처명', '부채_금액', '부채_계정']
    merged = pd.merge(grp_asset, grp_liab, on='거래처명', how='inner').sort_values(by='자산_금액', ascending=False)
    if not merged.empty:
        pd.DataFrame({'제목': ["자산/부채 동시 발생 거래처"]}).to_excel(writer, sheet_name='9_자산부채_교차', startrow=0, index=False, header=False)
        merged.to_excel(writer, sheet_name='9_자산부채_교차', startrow=1, index=False)

def run_revenue_expense_cross_check(df, writer, revenues, expenses):
    sheet_name = '10_매출비용_교차'
    if not revenues or not expenses:
        pd.DataFrame({'안내': ['매출 계정 또는 비용 계정을 입력하세요.']}).to_excel(writer, sheet_name=sheet_name, index=False)
        return
    print("   ▶ [5. 교차분석] 매출 vs 비용")
    rev_mask = pd.Series([False] * len(df))
    for r in revenues: rev_mask |= df[COL_ACCOUNT].str.contains(r, na=False, regex=False)
    rev_df = df[rev_mask & (df[COL_CREDIT] > 0)]
    exp_mask = pd.Series([False] * len(df))
    for e in expenses: exp_mask |= df[COL_ACCOUNT].str.contains(e, na=False, regex=False)
    exp_df = df[exp_mask & (df[COL_DEBIT] > 0)]
    if rev_df.empty or exp_df.empty:
        pd.DataFrame({'안내': ['해당 조건의 매출 또는 비용 데이터가 없습니다.']}).to_excel(writer, sheet_name=sheet_name, index=False)
        return
    grp_rev = rev_df.groupby(COL_CLIENT).agg({COL_CREDIT: 'sum', COL_ACCOUNT: lambda x: ','.join(set(x))}).reset_index()
    grp_rev.columns = ['거래처명', '매출_금액', '매출_계정']
    grp_exp = exp_df.groupby(COL_CLIENT).agg({COL_DEBIT: 'sum', COL_ACCOUNT: lambda x: ','.join(set(x))}).reset_index()
    grp_exp.columns = ['거래처명', '비용_금액', '비용_계정']
    merged = pd.merge(grp_rev, grp_exp, on='거래처명', how='inner').sort_values(by='매출_금액', ascending=False)
    if not merged.empty:
        pd.DataFrame({'제목': ["매출/비용 동시 발생 거래처"]}).to_excel(writer, sheet_name=sheet_name, startrow=0, index=False, header=False)
        merged.to_excel(writer, sheet_name=sheet_name, startrow=1, index=False)
    else:
        pd.DataFrame({'안내': ['매출·비용 동시 발생 거래처가 없습니다.']}).to_excel(writer, sheet_name=sheet_name, index=False)

def _get_구분_column(df):
    """구분 컬럼명 반환 (공백 등 차이 무시). 없으면 None"""
    for c in df.columns:
        if str(c).strip() == '구분':
            return c
    return None

def run_user_defined_top_analysis(df, writer, config, period_label=None):
    """config: list of (acct_name, top_n, direction). direction = '차변' | '대변' | 'both'
    df는 메인에서 기간 필터된 df_work가 전달될 수 있음. period_label로 구분 컬럼 없을 때 보강."""
    if not config: return
    work = df.copy()
    period_label = period_label or '당기+전기'
    gubun_col = _get_구분_column(work)
    if gubun_col is None and period_label in ('당기만', '전기만'):
        gubun_val = '당기' if period_label == '당기만' else '전기'
        work['구분'] = gubun_val
        gubun_col = '구분'
    has_구분 = gubun_col is not None
    if has_구분:
        n_gubun = work[gubun_col].nunique()
        if n_gubun == 1:
            period_label = str(work[gubun_col].iloc[0]) + '만'
    print(f"   ▶ [6. 심층분석] {len(config)}개 계정 자동 분석 중... (기간: {period_label})")
    for idx, (acct_name, top_n, direction) in enumerate(config, start=1):
        filtered = work[work[COL_ACCOUNT].str.contains(acct_name, na=False, regex=False)]
        if filtered.empty: continue
        safe_name = re.sub(r'[\\/*?:\[\]]', '', acct_name)[:18]
        sheet_name = f'Top_{idx}_{safe_name}'[:31]
        show_debit = direction in ('차변', 'both')
        show_credit = direction in ('대변', 'both')

        group_cols = [COL_CLIENT]
        if has_구분 and gubun_col in filtered.columns:
            group_cols = [gubun_col, COL_CLIENT]

        debit_top = pd.DataFrame()
        if show_debit:
            debit_rows = filtered[filtered[COL_DEBIT] > 0]
            if not debit_rows.empty and COL_CLIENT in df.columns:
                debit_top = debit_rows.groupby(group_cols)[COL_DEBIT].agg(['count', 'sum']).reset_index()
                debit_top = debit_top.sort_values(by='sum', ascending=False).head(top_n * 2 if has_구분 else top_n)
                if has_구분:
                    if gubun_col in debit_top.columns:
                        debit_top = debit_top.rename(columns={gubun_col: '구분'})
                        debit_top = debit_top[['구분', COL_CLIENT, 'count', 'sum']]
                        debit_top.columns = ['구분', '차변거래처명', '전표개수(차)', '차변금액']
                    else:
                        debit_top.columns = [COL_CLIENT, 'count', 'sum']
                        debit_top.columns = ['차변거래처명', '전표개수(차)', '차변금액']
                        gubun_val = debit_rows[gubun_col].iloc[0] if gubun_col in debit_rows.columns else ''
                        debit_top.insert(0, '구분', gubun_val)
                else:
                    debit_top.columns = ['차변거래처명', '전표개수(차)', '차변금액']
                debit_top.insert(0, '계정명', acct_name)
                debit_top = debit_top.reset_index(drop=True)

        credit_top = pd.DataFrame()
        if show_credit:
            credit_rows = filtered[filtered[COL_CREDIT] > 0]
            if not credit_rows.empty and COL_CLIENT in df.columns:
                credit_top = credit_rows.groupby(group_cols)[COL_CREDIT].agg(['count', 'sum']).reset_index()
                credit_top = credit_top.sort_values(by='sum', ascending=False).head(top_n * 2 if has_구분 else top_n)
                if has_구분:
                    if gubun_col in credit_top.columns:
                        credit_top = credit_top.rename(columns={gubun_col: '구분'})
                        credit_top = credit_top[['구분', COL_CLIENT, 'count', 'sum']]
                        credit_top.columns = ['구분', '대변거래처명', '전표개수(대)', '대변금액']
                    else:
                        credit_top.columns = [COL_CLIENT, 'count', 'sum']
                        credit_top.columns = ['대변거래처명', '전표개수(대)', '대변금액']
                        gubun_val = credit_rows[gubun_col].iloc[0] if gubun_col in credit_rows.columns else ''
                        credit_top.insert(0, '구분', gubun_val)
                else:
                    credit_top.columns = ['대변거래처명', '전표개수(대)', '대변금액']
                credit_top.insert(0, '계정명', acct_name)
                credit_top = credit_top.reset_index(drop=True)

        if show_debit and show_credit:
            combined = pd.concat([debit_top, credit_top], axis=1)
        elif show_debit:
            combined = debit_top
        else:
            combined = credit_top

        dir_label = '차변·대변' if direction == 'both' else direction
        if not combined.empty:
            title = f"[{acct_name}] 상위 {top_n}개 거래처 분석 ({dir_label}, {period_label})"
            pd.DataFrame({'제목': [title]}).to_excel(writer, sheet_name=sheet_name, startrow=0, index=False, header=False)
            combined.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)

# [★수정됨] 시트 이름 한글로 변경
def run_counterpart_analysis(df, writer, target_acct, target_dir):
    print(f"\n   🔍 [8. 상대계정] 목표: '{target_acct}' ({target_dir}) 분석 시작")
    if not target_acct: 
        print("     ⚠️ [Pass] 입력된 계정명이 없습니다. (건너뜀)")
        return

    val_col = COL_DEBIT if target_dir == '차변' else COL_CREDIT
    mask = (df[COL_ACCOUNT].astype(str).str.contains(target_acct, na=False)) & (df[val_col] > 0)
    target_rows = df[mask]
    
    count = len(target_rows)
    print(f"     👉 1단계: '{target_acct}'가 포함된 '{target_dir}' 전표 {count}건 발견")
    
    if target_rows.empty:
        print(f"     ⚠️ [실패] 해당 조건의 데이터가 없습니다. (계정명 오타 확인 필요)")
        return

    target_jids = target_rows[COL_JOURNAL_ID].unique()
    print(f"     👉 2단계: 연관된 전표번호 {len(target_jids)}개 추출 완료")

    related_rows = df[df[COL_JOURNAL_ID].isin(target_jids)].copy()
    print(f"     👉 3단계: 전체 전표 데이터에서 {len(related_rows)}행 로드함")

    counterpart_summary = related_rows.groupby(COL_ACCOUNT)[[COL_DEBIT, COL_CREDIT]].agg(['sum', 'count']).reset_index()
    counterpart_summary.columns = ['상대계정명', '차변합계', '차변건수', '대변합계', '대변건수']
    
    sort_col = '대변합계' if target_dir == '차변' else '차변합계'
    counterpart_summary = counterpart_summary.sort_values(by=sort_col, ascending=False)

    # [수정] 시트명 한글로 변경
    safe_name = re.sub(r"[^가-힣a-zA-Z0-9]", "", target_acct)[:15]
    sheet_name = f'상대계정분석_{safe_name}'  # 예: 상대계정분석_접대비
    
    print(f"     👉 4단계: 엑셀 시트 '{sheet_name}' 생성 및 저장 중...")
    
    pd.DataFrame({'A': [f"분석대상: {target_acct} ({target_dir})", f"전표 수: {len(target_jids)}"]}).to_excel(writer, sheet_name=sheet_name, index=False, header=False)
    counterpart_summary.to_excel(writer, sheet_name=sheet_name, startrow=3, index=False)
    print("     ✅ [완료] 상대계정 분석 성공!")

def run_ai_preparation(df, target_acct, output_writer):
    if not target_acct: return
    print(f"   ▶ [7. AI준비] '{target_acct}' 처리 중...")
    filtered = df[df[COL_ACCOUNT].str.contains(target_acct, na=False, regex=False)].copy()
    if filtered.empty: return
    safe_name = re.sub(r'[\\/*?:\[\]]', '', target_acct)[:10]
    sheet_name = f'AI_{safe_name}'[:30]
    filtered['YM'] = pd.to_datetime(filtered[COL_DATE], errors='coerce').dt.strftime('%Y-%m')
    monthly = filtered.groupby('YM')[[COL_DEBIT, COL_CREDIT]].agg(['sum', 'count']).reset_index()
    monthly.columns = ['YM', '차변합계', '차변건수', '대변합계', '대변건수']
    filtered['MaxAmt'] = filtered[[COL_DEBIT, COL_CREDIT]].max(axis=1)
    if len(filtered) > 10:
        cutoff = filtered['MaxAmt'].quantile(0.90)
        sample = filtered[filtered['MaxAmt'] >= cutoff].copy()
    else: sample = filtered.copy()
    if COL_CLIENT in sample.columns: sample[COL_CLIENT] = sample[COL_CLIENT].apply(get_safe_client_name)
    for col in MASK_TARGET_COLS:
        if col in sample.columns: sample[col] = sample[col].apply(mask_sensitive_info)
    sample = sample.drop(columns=['MaxAmt', 'YM'], errors='ignore')
    row = 0
    pd.DataFrame({'A': ["1. 월별 통계"]}).to_excel(output_writer, sheet_name=sheet_name, startrow=row, index=False, header=False)
    monthly.to_excel(output_writer, sheet_name=sheet_name, startrow=row+1, index=False)
    row += len(monthly) + 5
    pd.DataFrame({'A': ["2. 상위 10% 샘플 (익명화 & 마스킹)"]}).to_excel(output_writer, sheet_name=sheet_name, startrow=row, index=False, header=False)
    sample.to_excel(output_writer, sheet_name=sheet_name, startrow=row+1, index=False)

def run_account_list(df, writer):
    """분개장에 사용된 모든 계정명을 리스트로 정리"""
    print("   ▶ [계정명리스트] 사용된 계정명 추출 중...")
    
    if COL_ACCOUNT not in df.columns:
        print("     ⚠️ 계정명 컬럼을 찾을 수 없습니다.")
        return
    
    # 사용된 모든 계정명 추출 (중복 제거)
    account_list = df[COL_ACCOUNT].dropna().unique()
    account_list = [str(acc).strip() for acc in account_list if str(acc).strip() != '']
    account_list = sorted(set(account_list))  # 정렬 및 중복 제거
    
    if not account_list:
        print("     ⚠️ 계정명 데이터가 없습니다.")
        return
    
    # 계정명별 통계 정보도 함께 제공
    account_stats = df.groupby(COL_ACCOUNT).agg({
        COL_DEBIT: ['sum', 'count'],
        COL_CREDIT: ['sum', 'count'],
        COL_JOURNAL_ID: 'nunique'
    }).reset_index()
    account_stats.columns = ['계정명', '차변합계', '차변건수', '대변합계', '대변건수', '전표개수']
    
    # 차변과 대변 중 큰 값 기준으로 정렬
    account_stats['최대금액'] = account_stats[['차변합계', '대변합계']].max(axis=1)
    account_stats = account_stats.sort_values(by='최대금액', ascending=False)
    account_stats = account_stats.drop(columns=['최대금액'])
    
    # 시트에 저장
    sheet_name = '계정명리스트'
    
    # 1. 간단한 리스트 (계정명만)
    simple_list = pd.DataFrame({'계정명': account_list})
    pd.DataFrame({'제목': ['[사용된 계정명 리스트]', f'총 {len(account_list)}개 계정']}).to_excel(
        writer, sheet_name=sheet_name, startrow=0, index=False, header=False
    )
    simple_list.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
    
    # 2. 통계 정보 포함 리스트
    start_row = len(simple_list) + 5
    pd.DataFrame({'제목': ['[계정명별 통계 정보]']}).to_excel(
        writer, sheet_name=sheet_name, startrow=start_row, index=False, header=False
    )
    account_stats.to_excel(writer, sheet_name=sheet_name, startrow=start_row+1, index=False)
    
    print(f"     ✅ 계정명 리스트 완료: {len(account_list)}개 계정")

def run_date_difference_analysis(df, writer, days_threshold):
    """전표일자와 등록일자 차이 분석"""
    if days_threshold is None or days_threshold <= 0:
        # 사용자가 입력하지 않았거나 0 이하 입력 시 시트 생성하지 않음
        return
    
    print(f"   ▶ [일자차이분석] 전표일자와 등록일자 차이 {days_threshold}일 이상인 전표 분석 중...")
    
    # 등록일자 컬럼 찾기
    reg_date_col = None
    possible_names = ['등록일자', '등록일', '작성일자', '작성일', '생성일자', '생성일', '입력일자', '입력일']
    for col in df.columns:
        col_str = str(col).strip()
        if any(name in col_str for name in possible_names):
            reg_date_col = col
            break
    
    if reg_date_col is None:
        print("     ⚠️ 등록일자 컬럼을 찾을 수 없습니다. 컬럼 목록:", df.columns.tolist())
        # 등록일자 컬럼을 찾지 못해도 시트는 생성하여 확인할 수 있도록 함
        sheet_name = '일자차이분석'
        pd.DataFrame({
            '오류': [
                f"기준 일자 차이: {days_threshold}일 이상",
                "오류: 등록일자 컬럼을 찾을 수 없습니다.",
                f"사용 가능한 컬럼: {', '.join(df.columns.tolist()[:10])}..."  # 처음 10개만 표시
            ]
        }).to_excel(writer, sheet_name=sheet_name, startrow=0, index=False, header=False)
        return
    
    # 전표일자와 등록일자 모두 있는 데이터만 필터링
    df_date = df[(df[COL_DATE].notna()) & (df[reg_date_col].notna())].copy()
    if df_date.empty:
        print("     ⚠️ 전표일자 또는 등록일자가 없는 데이터입니다.")
        # 데이터가 없어도 시트는 생성하여 확인할 수 있도록 함
        sheet_name = '일자차이분석'
        pd.DataFrame({
            '정보': [
                f"기준 일자 차이: {days_threshold}일 이상",
                "오류: 전표일자 또는 등록일자가 없는 데이터입니다.",
                "데이터를 확인해주세요."
            ]
        }).to_excel(writer, sheet_name=sheet_name, startrow=0, index=False, header=False)
        return
    
    # 등록일자도 날짜 형식으로 변환 시도
    if df_date[reg_date_col].dtype == 'object' or df_date[reg_date_col].dtype == 'int64':
        # 숫자 형식일 경우 (예: 20250101)
        try:
            df_date[reg_date_col] = pd.to_numeric(df_date[reg_date_col], errors='coerce')
            df_date[reg_date_col] = df_date[reg_date_col].fillna(0).astype('int64').astype(str)
            df_date[reg_date_col] = pd.to_datetime(df_date[reg_date_col], format='%Y%m%d', errors='coerce')
        except:
            # 이미 날짜 형식이거나 다른 형식일 경우
            df_date[reg_date_col] = pd.to_datetime(df_date[reg_date_col], errors='coerce')
    else:
        df_date[reg_date_col] = pd.to_datetime(df_date[reg_date_col], errors='coerce')
    
    # 날짜 차이 계산 (등록일자 - 전표일자)
    df_date['일자차이'] = (df_date[reg_date_col] - df_date[COL_DATE]).dt.days
    
    # 사용자가 입력한 일자 이상 차이나는 전표 필터링
    filtered = df_date[df_date['일자차이'] >= days_threshold].copy()
    
    if filtered.empty:
        print(f"     ℹ️ {days_threshold}일 이상 차이나는 전표가 없습니다.")
        # 조건에 맞는 데이터가 없어도 시트는 생성하여 확인할 수 있도록 함
        sheet_name = '일자차이분석'
        pd.DataFrame({
            '정보': [
                f"기준 일자 차이: {days_threshold}일 이상",
                f"발견된 전표 수: 0개",
                f"총 행 수: 0행",
                f"결과: 조건에 맞는 전표가 없습니다."
            ]
        }).to_excel(writer, sheet_name=sheet_name, startrow=0, index=False, header=False)
        return
    
    # 전표번호별로 그룹화하여 집계
    journal_summary = filtered.groupby(COL_JOURNAL_ID).agg({
        COL_DATE: 'first',
        reg_date_col: 'first',
        '일자차이': 'first',
        COL_ACCOUNT: lambda x: ', '.join(x.unique()[:5]),  # 계정명 (최대 5개)
        COL_DEBIT: 'sum',
        COL_CREDIT: 'sum',
        COL_CLIENT: lambda x: ', '.join(x.dropna().unique()[:3]) if COL_CLIENT in filtered.columns else ''  # 거래처명 (최대 3개)
    }).reset_index()
    
    journal_summary.columns = ['전표번호', '전표일자', '등록일자', '일자차이', '계정명', '차변합계', '대변합계', '거래처명']
    
    # 일자차이 내림차순 정렬
    journal_summary = journal_summary.sort_values(by='일자차이', ascending=False)
    
    # 상세 내역도 준비
    detail_df = filtered.copy()
    detail_df = detail_df.sort_values(by=['일자차이', COL_JOURNAL_ID], ascending=[False, True])
    
    # 시트에 저장
    sheet_name = '일자차이분석'
    
    # 요약 정보
    info_rows = [
        f"기준 일자 차이: {days_threshold}일 이상",
        f"발견된 전표 수: {journal_summary['전표번호'].nunique()}개",
        f"총 행 수: {len(filtered)}행",
        f"최대 일자 차이: {journal_summary['일자차이'].max()}일"
    ]
    pd.DataFrame({'정보': info_rows}).to_excel(writer, sheet_name=sheet_name, startrow=0, index=False, header=False)
    
    # 전표별 요약
    pd.DataFrame({'제목': ['[전표별 요약]']}).to_excel(writer, sheet_name=sheet_name, startrow=len(info_rows)+2, index=False, header=False)
    journal_summary.to_excel(writer, sheet_name=sheet_name, startrow=len(info_rows)+3, index=False)
    
    # 상세 내역
    start_row = len(info_rows) + len(journal_summary) + 6
    pd.DataFrame({'제목': ['[상세 내역]']}).to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False, header=False)
    
    # 상세 내역에서 필요한 컬럼만 선택
    detail_cols = [COL_JOURNAL_ID, COL_DATE, reg_date_col, '일자차이', COL_ACCOUNT, COL_DEBIT, COL_CREDIT]
    if COL_CLIENT in detail_df.columns:
        detail_cols.append(COL_CLIENT)
    if COL_DESC in detail_df.columns:
        detail_cols.append(COL_DESC)
    
    detail_output = detail_df[detail_cols].copy()
    detail_output.columns = [col if col != reg_date_col else '등록일자' for col in detail_output.columns]
    detail_output.to_excel(writer, sheet_name=sheet_name, startrow=start_row+1, index=False)
    
    print(f"     ✅ 일자차이 분석 완료: {len(journal_summary)}개 전표, {len(filtered)}행")

def run_employee_summary(df, writer):
    """사원명 기준으로 차변/대변 계정별 집계"""
    print("   ▶ [사원별 집계] 사원명 기준 데이터 정리 중...")
    
    # 사원명 컬럼 찾기
    employee_col = None
    possible_names = ['사원명', '작성자', '작성자명', '사용자', 'User', 'Employee']
    for col in df.columns:
        col_str = str(col).strip()
        if any(name in col_str for name in possible_names):
            employee_col = col
            break
    
    if employee_col is None:
        print("     ⚠️ 사원명 컬럼을 찾을 수 없습니다. 컬럼 목록:", df.columns.tolist())
        return
    
    # 사원명이 없는 행 처리
    df_emp = df[df[employee_col].notna() & (df[employee_col].astype(str).str.strip() != '')].copy()
    if df_emp.empty:
        print("     ⚠️ 사원명 데이터가 없습니다.")
        return
    
    # 차변 데이터 집계 (사원명, 계정명별)
    debit_data = df_emp[df_emp[COL_DEBIT] > 0].copy()
    if not debit_data.empty:
        debit_summary = debit_data.groupby([employee_col, COL_ACCOUNT]).agg({
            COL_DEBIT: 'sum',
            COL_JOURNAL_ID: 'nunique'
        }).reset_index()
        debit_summary.columns = ['사원명', '차변계정명', '차변금액', '차변전표개수']
    else:
        debit_summary = pd.DataFrame(columns=['사원명', '차변계정명', '차변금액', '차변전표개수'])
    
    # 대변 데이터 집계 (사원명, 계정명별)
    credit_data = df_emp[df_emp[COL_CREDIT] > 0].copy()
    if not credit_data.empty:
        credit_summary = credit_data.groupby([employee_col, COL_ACCOUNT]).agg({
            COL_CREDIT: 'sum',
            COL_JOURNAL_ID: 'nunique'
        }).reset_index()
        credit_summary.columns = ['사원명', '대변계정명', '대변금액', '대변전표개수']
    else:
        credit_summary = pd.DataFrame(columns=['사원명', '대변계정명', '대변금액', '대변전표개수'])
    
    # 사원명별로 차변과 대변을 조인
    # 각 사원의 모든 차변계정과 대변계정을 조합
    result_rows = []
    
    # 모든 사원명 가져오기
    all_employees = set()
    if not debit_summary.empty:
        all_employees.update(debit_summary['사원명'].unique())
    if not credit_summary.empty:
        all_employees.update(credit_summary['사원명'].unique())
    
    for emp in sorted(all_employees):
        emp_debit = debit_summary[debit_summary['사원명'] == emp] if not debit_summary.empty else pd.DataFrame()
        emp_credit = credit_summary[credit_summary['사원명'] == emp] if not credit_summary.empty else pd.DataFrame()
        
        # 차변과 대변의 최대 행 수 계산
        max_rows = max(len(emp_debit), len(emp_credit), 1)
        
        for i in range(max_rows):
            row = {'사원명': emp if i == 0 else ''}
            
            # 차변 정보
            if i < len(emp_debit):
                row['차변계정명'] = emp_debit.iloc[i]['차변계정명']
                row['차변금액'] = emp_debit.iloc[i]['차변금액']
                row['전표개수'] = emp_debit.iloc[i]['차변전표개수']  # 차변 전표개수
            else:
                row['차변계정명'] = ''
                row['차변금액'] = 0
                row['전표개수'] = 0
            
            # 대변 정보
            if i < len(emp_credit):
                row['대변계정명'] = emp_credit.iloc[i]['대변계정명']
                row['대변금액'] = emp_credit.iloc[i]['대변금액']
                row['대변전표개수'] = emp_credit.iloc[i]['대변전표개수']  # 대변 전표개수
            else:
                row['대변계정명'] = ''
                row['대변금액'] = 0
                row['대변전표개수'] = 0
            
            result_rows.append(row)
    
    result_df = pd.DataFrame(result_rows)
    
    # 컬럼 순서 정리: 사원명, 차변계정명, 차변금액, 전표개수, 대변계정명, 대변금액, 전표개수
    if not result_df.empty:
        # 컬럼 순서: 사원명, 차변계정명, 차변금액, 전표개수(차변), 대변계정명, 대변금액, 전표개수(대변)
        # 사용자 요청 형식에 맞게 컬럼명 조정
        # pandas에서는 중복 컬럼명을 피하기 위해 임시로 다른 이름 사용
        result_df = result_df[['사원명', '차변계정명', '차변금액', '전표개수', '대변계정명', '대변금액', '대변전표개수']]
        # Excel에 저장할 때 컬럼명을 사용자 요청 형식으로 변경
        result_df.columns = ['사원명', '차변계정명', '차변금액', '전표개수', '대변계정명', '대변금액', '전표개수']
        result_df.to_excel(writer, sheet_name='사원별집계', index=False)
        print(f"     ✅ 사원별 집계 완료: {len(result_df)}행 생성")
    else:
        print("     ⚠️ 집계 결과가 없습니다.")

def analyze_client_comparison(df, account_list, value_type='차변'):
    """
    계정별 거래처 전기/당기 비교 분석.
    - account_list: 계정과목명 리스트 (쉼표로 복수 입력 가능, 부분 일치)
    - value_type: '차변', '대변', 또는 'both' (차변+대변 합산)
    반환: 계정명 | 거래처명 | 전기금액 | 당기금액 | 전기전표수 | 당기전표수 | 증감금액 | 증감비율(%)
    """
    if COL_ACCOUNT not in df.columns or COL_CLIENT not in df.columns:
        return pd.DataFrame()
    if '구분' not in df.columns:
        return pd.DataFrame()
    if isinstance(account_list, str):
        account_list = [a.strip() for a in account_list.split(',') if a.strip()]
    if not account_list:
        return pd.DataFrame()

    # 구분 컬럼 값 정규화 (공백 제거) → 피벗 컬럼이 '전기'/'당기'로 통일되어 전기금액/당기금액 컬럼 생성됨
    df_work = df.copy()
    gubun_col = next((c for c in df_work.columns if str(c).strip() == '구분'), None)
    if gubun_col is not None:
        df_work[gubun_col] = df_work[gubun_col].astype(str).str.strip()

    # 1. 데이터 필터링: 선택한 계정과목들 (OR 조건, 괄호 전각/반각·공백 차이 무시)
    mask = pd.Series(False, index=df_work.index)
    for acct in account_list:
        mask = mask | _account_match_flexible(df_work[COL_ACCOUNT], acct)
    filtered = df_work[mask].copy()
    if filtered.empty:
        return pd.DataFrame()

    # 금액 컬럼 결정 (피벗용 단일 컬럼 '_amt'로 통일)
    if value_type == '차변':
        filtered['_amt'] = pd.to_numeric(filtered[COL_DEBIT], errors='coerce').fillna(0)
    elif value_type == '대변':
        filtered['_amt'] = pd.to_numeric(filtered[COL_CREDIT], errors='coerce').fillna(0)
    elif value_type == 'both':
        filtered['_amt'] = pd.to_numeric(filtered[COL_DEBIT], errors='coerce').fillna(0) + pd.to_numeric(filtered[COL_CREDIT], errors='coerce').fillna(0)
    else:
        filtered['_amt'] = pd.to_numeric(filtered[COL_DEBIT], errors='coerce').fillna(0)
    amt_col = '_amt'

    # 2. 피벗 테이블: 인덱스=(계정명, 거래처명), 컬럼=구분, 값=금액(sum) + 전표수(count)
    pivot = filtered.pivot_table(
        index=[COL_ACCOUNT, COL_CLIENT],
        columns='구분',
        values=amt_col,
        aggfunc=['sum', 'count'],
        fill_value=0
    )

    if pivot.empty:
        return pd.DataFrame()

    # 3. 컬럼 평탄화: 전기금액, 당기금액, 전기전표수, 당기전표수
    result = pd.DataFrame(index=pivot.index)
    for col in ['전기금액', '당기금액', '전기전표수', '당기전표수']:
        result[col] = 0
    for (agg_name, gubun) in pivot.columns:
        g = str(gubun).strip()  # '전기 '/'당기 ' 등 공백 제거 → 전기금액/당기금액 컬럼명 통일
        if agg_name == 'sum':
            result[f'{g}금액'] = pivot[(agg_name, gubun)].reindex(result.index).fillna(0)
        else:
            result[f'{g}전표수'] = pivot[(agg_name, gubun)].reindex(result.index).fillna(0).astype(int)

    result = result.fillna(0)
    for c in ['전기전표수', '당기전표수']:
        result[c] = result[c].astype(int)

    # 4. 파생 컬럼 (전기금액 0이면 ZeroDivisionError 방지)
    result['증감금액'] = result['당기금액'] - result['전기금액']
    result['증감비율(%)'] = result.apply(
        lambda r: (r['증감금액'] / r['전기금액'] * 100) if r['전기금액'] != 0 else 0.0,
        axis=1
    )

    # 5. 정렬: 계정명 → 증감금액 절댓값 큰 순 → 당기금액 큰 순
    result['_abs_change'] = result['증감금액'].abs()
    result = result.sort_values(by=[COL_ACCOUNT, '_abs_change', '당기금액'], ascending=[True, False, False])
    result = result.drop(columns=['_abs_change'])

    result = result.reset_index()
    # pivot 인덱스 → 계정명/거래처명 (일부 환경에서 level_0/level_1로 나올 수 있음)
    cols = list(result.columns)
    if len(cols) >= 2 and ('계정명' not in cols or '거래처명' not in cols):
        renames = {}
        if cols[0] != '계정명':
            renames[cols[0]] = '계정명'
        if cols[1] != '거래처명':
            renames[cols[1]] = '거래처명'
        if renames:
            result = result.rename(columns=renames)
    # 계정명/거래처명 앞뒤 공백 제거 → 시트별 필터 시 빈 시트 방지 (미수대매금 vs 미수대매금 )
    if '계정명' in result.columns:
        result['계정명'] = result['계정명'].astype(str).str.strip()
    if '거래처명' in result.columns:
        result['거래처명'] = result['거래처명'].astype(str).str.strip()
    out_cols = ['계정명', '거래처명', '전기금액', '당기금액', '전기전표수', '당기전표수', '증감금액', '증감비율(%)']
    result = result[[c for c in out_cols if c in result.columns]]
    return result

def analyze_client_comparison_top10_monthly(df, account_list, value_type):
    """
    선택된 계정·차/대변 기준 상위 10개 거래처의 월별 발생금액을 전기/당기 구분하여 반환.
    반환: (df_전기_월별, df_당기_월별) 각각 계정명 | 거래처명 | YM컬럼들 | 합계
    """
    if COL_ACCOUNT not in df.columns or COL_CLIENT not in df.columns or COL_DATE not in df.columns:
        return pd.DataFrame(), pd.DataFrame()
    if '구분' not in df.columns:
        return pd.DataFrame(), pd.DataFrame()
    if isinstance(account_list, str):
        account_list = [a.strip() for a in account_list.split(',') if a.strip()]
    if not account_list:
        return pd.DataFrame(), pd.DataFrame()

    # 구분 값 정규화 (공백 제거) → '전기 '/'당기 ' 비교 정상화
    df_work = df.copy()
    gubun_col = next((c for c in df_work.columns if str(c).strip() == '구분'), None)
    if gubun_col is not None:
        df_work[gubun_col] = df_work[gubun_col].astype(str).str.strip()

    mask = pd.Series(False, index=df_work.index)
    for acct in account_list:
        mask = mask | _account_match_flexible(df_work[COL_ACCOUNT], acct)
    filtered = df_work[mask].copy()
    if filtered.empty:
        return pd.DataFrame(), pd.DataFrame()

    if value_type == '차변':
        filtered['_amt'] = pd.to_numeric(filtered[COL_DEBIT], errors='coerce').fillna(0)
    elif value_type == '대변':
        filtered['_amt'] = pd.to_numeric(filtered[COL_CREDIT], errors='coerce').fillna(0)
    elif value_type == 'both':
        filtered['_amt'] = pd.to_numeric(filtered[COL_DEBIT], errors='coerce').fillna(0) + pd.to_numeric(filtered[COL_CREDIT], errors='coerce').fillna(0)
    else:
        filtered['_amt'] = pd.to_numeric(filtered[COL_DEBIT], errors='coerce').fillna(0)

    filtered['YM'] = pd.to_datetime(filtered[COL_DATE], errors='coerce').dt.strftime('%Y-%m')
    filtered = filtered[filtered['YM'].notna()]

    sub_전기 = filtered[filtered['구분'] == '전기']
    sub_당기 = filtered[filtered['구분'] == '당기']

    # 전기·당기 각각 금액 기준 상위 10개 거래처 (당기와 동일하게 거래처 월별 추출)
    top10_전기 = sub_전기.groupby([COL_ACCOUNT, COL_CLIENT])['_amt'].sum().groupby(level=1).sum().nlargest(10).index.tolist() if not sub_전기.empty else []
    top10_당기 = sub_당기.groupby([COL_ACCOUNT, COL_CLIENT])['_amt'].sum().groupby(level=1).sum().nlargest(10).index.tolist() if not sub_당기.empty else []

    def _pivot_monthly(sub, top_list):
        """해당 기간(전기 또는 당기) 상위 거래처 목록 기준으로 (계정명, 거래처명)별 월별 피벗 생성"""
        if not top_list:
            return pd.DataFrame()
        if sub.empty:
            return pd.DataFrame({'계정명': [], '거래처명': top_list, '합계': [0] * len(top_list)})
        pt = sub.pivot_table(index=[COL_ACCOUNT, COL_CLIENT], columns='YM', values='_amt', aggfunc='sum', fill_value=0)
        pt = pt.loc[pt.index.get_level_values(COL_CLIENT).isin(top_list)]
        pt['합계'] = pt.sum(axis=1)
        pt = pt.reset_index().rename(columns={COL_ACCOUNT: '계정명', COL_CLIENT: '거래처명'})
        pt = pt.sort_values(by=['계정명', '합계'], ascending=[True, False])
        return pt

    df_전기 = _pivot_monthly(sub_전기, top10_전기)
    df_당기 = _pivot_monthly(sub_당기, top10_당기)
    return df_전기, df_당기

def save_client_comparison_to_excel(df, account_list, value_type, base_dir=None, output_filename=None):
    """
    analyze_client_comparison 결과를 계정별로 별도 시트에 저장.
    각 시트: 거래처 전기/당기 비교표 + 전기/당기 월별 발생금액(상위 10개 거래처)
    account_list: 계정과목 문자열 또는 쉼표 구분 문자열 또는 리스트
    """
    target_path = _target_path(base_dir, filename=output_filename)
    result_df = analyze_client_comparison(df, account_list, value_type)
    df_전기_월별, df_당기_월별 = analyze_client_comparison_top10_monthly(df, account_list, value_type)

    def _write_sheets(writer):
        # 계정별 시트 분리: 계정명 컬럼이 있고, 안내용이 아닐 때만
        has_계정명 = '계정명' in result_df.columns and not result_df.empty
        if not has_계정명:
            result_df_use = result_df if not result_df.empty else pd.DataFrame({'안내': ['해당 조건의 거래처 비교 데이터가 없습니다.']})
            result_df_use.to_excel(writer, sheet_name='거래처_전기당기비교', index=False)
            if not df_전기_월별.empty or not df_당기_월별.empty:
                start_row = 0
                pd.DataFrame([['※ 전기 월별 발생금액 (상위 10개 거래처)', '']]).to_excel(writer, sheet_name='거래처_전기당기비교_월별', startrow=start_row, index=False, header=False)
                start_row += 2
                if not df_전기_월별.empty:
                    df_전기_월별.to_excel(writer, sheet_name='거래처_전기당기비교_월별', startrow=start_row, index=False)
                    start_row += len(df_전기_월별) + 2
                pd.DataFrame([['※ 당기 월별 발생금액 (상위 10개 거래처)', '']]).to_excel(writer, sheet_name='거래처_전기당기비교_월별', startrow=start_row, index=False, header=False)
                start_row += 2
                if not df_당기_월별.empty:
                    df_당기_월별.to_excel(writer, sheet_name='거래처_전기당기비교_월별', startrow=start_row, index=False)
            return

        # 계정별로 시트 분리
        accounts = result_df['계정명'].unique().tolist()
        used_names = set()
        for acct in accounts:
            base_name = _safe_sheet_name(f"비교_{acct}")
            sheet_name = base_name
            idx = 1
            while sheet_name in used_names:
                sheet_name = _safe_sheet_name(f"비교_{acct}_{idx}", max_len=29)
                idx += 1
            used_names.add(sheet_name)
            sub = result_df[result_df['계정명'] == acct].drop(columns=['계정명'])
            if sub.empty:
                pd.DataFrame({'안내': [f'계정 [{acct}] 에 해당하는 거래처 비교 데이터가 없습니다.']}).to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                sub.to_excel(writer, sheet_name=sheet_name, index=False)
            row = len(sub) + 2
            pd.DataFrame([['※ 전기 월별 발생금액 (상위 10개 거래처)', '']]).to_excel(writer, sheet_name=sheet_name, startrow=row, index=False, header=False)
            row += 2
            m_전기 = df_전기_월별[df_전기_월별['계정명'] == acct].drop(columns=['계정명']) if not df_전기_월별.empty and '계정명' in df_전기_월별.columns else pd.DataFrame()
            if not m_전기.empty:
                m_전기.to_excel(writer, sheet_name=sheet_name, startrow=row, index=False)
                row += len(m_전기) + 2
            else:
                pd.DataFrame([['(해당 기간 데이터 없음)', '']]).to_excel(writer, sheet_name=sheet_name, startrow=row, index=False, header=False)
                row += 2
            pd.DataFrame([['※ 당기 월별 발생금액 (상위 10개 거래처)', '']]).to_excel(writer, sheet_name=sheet_name, startrow=row, index=False, header=False)
            row += 2
            m_당기 = df_당기_월별[df_당기_월별['계정명'] == acct].drop(columns=['계정명']) if not df_당기_월별.empty and '계정명' in df_당기_월별.columns else pd.DataFrame()
            if not m_당기.empty:
                m_당기.to_excel(writer, sheet_name=sheet_name, startrow=row, index=False)
            else:
                pd.DataFrame([['(해당 기간 데이터 없음)', '']]).to_excel(writer, sheet_name=sheet_name, startrow=row, index=False, header=False)

    if os.path.exists(target_path):
        with pd.ExcelWriter(target_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            _write_sheets(writer)
    else:
        with pd.ExcelWriter(target_path, engine='openpyxl', mode='w') as writer:
            _write_sheets(writer)

    fname = os.path.basename(target_path)
    if result_df.empty or '계정명' not in result_df.columns:
        print(f"{fname} 파일에 저장되었습니다.")
    else:
        n = result_df['계정명'].nunique()
        print(f"{fname} 파일에 계정별 {n}개 시트로 저장되었습니다.")

# =============================================================================
# 🚀 [메인 실행부] - 대화형 메뉴 시스템
# =============================================================================

# 결과 저장 대상 파일 (메뉴 선택 시 mode='a' 로 누적 저장)
OUT_DIR_NAME = '분석결과_모음'
TARGET_EXCEL_NAME = 'JET_통합분석결과.xlsx'

def _target_path(base_dir=None, filename=None):
    if base_dir is None:
        base_dir = SCRIPT_DIR
    out_dir = os.path.normpath(os.path.join(base_dir, OUT_DIR_NAME.strip()))
    os.makedirs(out_dir, exist_ok=True)
    name = (filename or TARGET_EXCEL_NAME).strip()
    return os.path.normpath(os.path.join(out_dir, name))

def _get_writer(base_dir=None, filename=None):
    path = _target_path(base_dir, filename=filename)
    if os.path.exists(path):
        return pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace')
    return pd.ExcelWriter(path, engine='openpyxl', mode='w')

def _menu_filename(key, menu_name):
    """특정 번호 분석 시 파일명: JET_통합분석_{번호}_{메뉴명}.xlsx (공백/특수문자 제거)"""
    safe_name = re.sub(r'[\s/()]+', '', menu_name)  # 공백, /, 괄호 제거
    return f"JET_통합분석_{key}_{safe_name}.xlsx"

def _safe_sheet_name(name, max_len=31):
    """엑셀 시트명에 허용되지 않는 문자 제거 후 길이 제한"""
    s = re.sub(r'[\\/*?:\[\]]', '', str(name).strip())
    return (s[:max_len - 2] + '..') if len(s) > max_len else s

def _normalize_date_journal_columns(df):
    """
    분개장/계정별원장 등에서 쓰는 컬럼명을 앱 기준으로 통일.
    - 일자, 전표일자 → 전표일자
    - 전표등록번호, 전표번호 → 전표번호
    """
    cols = [str(c).strip() for c in df.columns]
    df.columns = cols
    if COL_DATE not in df.columns:
        for c in df.columns:
            if c in ('일자', '전표일자', '전표일', '거래일자', '적요일자'):
                df = df.rename(columns={c: COL_DATE})
                break
    if COL_JOURNAL_ID not in df.columns:
        for c in df.columns:
            if c in ('전표등록번호', '전표번호', '전표NO', '전표 no'):
                df = df.rename(columns={c: COL_JOURNAL_ID})
                break
    if COL_CLIENT not in df.columns:
        for c in df.columns:
            if c in ('거래처명', '거래처', '상대거래처', '거래처명칭'):
                df = df.rename(columns={c: COL_CLIENT})
                break
    if COL_ACCOUNT not in df.columns:
        for c in df.columns:
            if c in ('계정명', '계정과목', '계정', '과목'):
                df = df.rename(columns={c: COL_ACCOUNT})
                break
    return df

def _normalize_debit_credit_columns(df):
    """
    실제 파일의 차변/대변 컬럼명(차변금액, 대변금액 등)을
    앱에서 사용하는 '차변', '대변'으로 매칭·통일.
    전기/당기 파일마다 컬럼명이 다를 수 있으므로, '차변' 포함 컬럼은 모두 하나로 합치고
    쉼표 포맷(예: 97,900)도 숫자로 변환.
    """
    cols = [str(c).strip() for c in df.columns]
    df.columns = cols
    # '차변' 포함 컬럼들 → 쉼표 제거 등 숫자 변환 후 하나로 합침 (전기/당기 컬럼명 상이 대비)
    debit_cols = [c for c in df.columns if '차변' in c]
    if debit_cols:
        for c in debit_cols:
            df[c] = _to_numeric_amount(df[c])
        primary = COL_DEBIT if COL_DEBIT in debit_cols else debit_cols[0]
        df[COL_DEBIT] = df[primary].copy()
        for c in debit_cols:
            if c != COL_DEBIT:
                df[COL_DEBIT] = df[COL_DEBIT].fillna(df[c])
                df = df.drop(columns=[c], errors='ignore')
    # '대변' 포함 컬럼들 → 숫자 변환 후 하나로 합침
    credit_cols = [c for c in df.columns if '대변' in c and c != COL_DEBIT]
    if credit_cols:
        for c in credit_cols:
            df[c] = _to_numeric_amount(df[c])
        primary = COL_CREDIT if COL_CREDIT in credit_cols else credit_cols[0]
        df[COL_CREDIT] = df[primary].copy()
        for c in credit_cols:
            if c != COL_CREDIT:
                df[COL_CREDIT] = df[COL_CREDIT].fillna(df[c])
                df = df.drop(columns=[c], errors='ignore')
    return df

def _to_numeric_amount(series):
    """쉼표·공백 제거 후 숫자 변환 (예: '97,900' -> 97900).
    엑셀에서 셀 너비 부족으로 표시된 '#######' 등 #만 있는 문자열은 0으로 처리."""
    s = series.astype(str).str.strip()
    s = s.str.replace(',', '', regex=False).str.replace(' ', '', regex=False)
    s = s.str.replace(r'^#+$', '0', regex=True)  # ####### 등 → 0
    return pd.to_numeric(s, errors='coerce').fillna(0)

def _preprocess_df(df):
    """공통 전처리: 컬럼 정리(일자/전표등록번호 등), 날짜/숫자 변환"""
    df = df.copy()
    df.columns = df.columns.str.strip()
    df = _normalize_date_journal_columns(df)
    df = _normalize_debit_credit_columns(df)
    if COL_DATE in df.columns:
        # YYYYMMDD(숫자/문자) 또는 YYYY-MM-DD 등 다양한 형식 허용
        ser = df[COL_DATE]
        try:
            num = pd.to_numeric(ser, errors='coerce')
            if num.notna().sum() > len(ser) * 0.5:
                ser = num.fillna(0).astype('int64').astype(str)
                df[COL_DATE] = pd.to_datetime(ser, format='%Y%m%d', errors='coerce')
            else:
                df[COL_DATE] = pd.to_datetime(ser, errors='coerce')
        except Exception:
            df[COL_DATE] = pd.to_datetime(ser, errors='coerce')
    if COL_DEBIT in df.columns:
        df[COL_DEBIT] = _to_numeric_amount(df[COL_DEBIT])
    if COL_CREDIT in df.columns:
        df[COL_CREDIT] = _to_numeric_amount(df[COL_CREDIT])
    possible_names = ['등록일자', '등록일', '작성일자', '작성일', '생성일자', '생성일', '입력일자', '입력일']
    for col in df.columns:
        if any(name in str(col).strip() for name in possible_names):
            try:
                df[col] = pd.to_numeric(df[col], errors='coerce')
                df[col] = df[col].fillna(0).astype('int64').astype(str)
                df[col] = pd.to_datetime(df[col], format='%Y%m%d', errors='coerce')
            except Exception:
                df[col] = pd.to_datetime(df[col], errors='coerce')
            break
    return df

def _has_구분_column(df):
    """컬럼명 공백 제거 후 '구분' 존재 여부"""
    return any(str(c).strip() == '구분' for c in df.columns)

def _ask_period_filter(df):
    """금액·건수 집계 시 분개장과 비교하려면 기간을 맞추면 됨. (당기+전기 합산이 기본)
    반환: (df_work, period_label) — df_work는 필터된 또는 원본 df, period_label은 '당기만'/'전기만'/'당기+전기'"""
    print("", flush=True)
    print("   ★★★ 기간 선택 ★★★", flush=True)
    print("   1 = 당기만   2 = 전기만   엔터 = 전체 (분개장과 비교 시 같은 기간 선택)", flush=True)
    if not _has_구분_column(df):
        print("   → 데이터에 당기/전기 구분이 없어 전체로 진행합니다.", flush=True)
        return df.copy(), '전체'
    p_in = input("   기간 입력 (1/2/엔터): ").strip()
    if p_in == '1':
        return df[df['구분'] == '당기'].copy(), '당기만'
    if p_in == '2':
        return df[df['구분'] == '전기'].copy(), '전기만'
    return df.copy(), '당기+전기'

# ---------- 메뉴별 실행 함수 (각각 JET_통합분석결과.xlsx 에 시트 추가) ----------
def run_menu_client_comparison(df, base_dir=None, output_filename=None):
    print("\n   [거래처 전기/당기 비교]")
    print("   계정과목: 쉼표로 여러 계정 입력 가능 (예: 접대비, 복리후생비)")
    account_in = input("   계정과목 (예: 접대비): ").strip() or '접대비'
    account_list = [a.strip() for a in account_in.split(',') if a.strip()]
    value_type = input("   금액 기준 (차변/대변/both=둘다, 엔터=차변): ").strip() or '차변'
    if value_type not in ('차변', '대변', 'both'):
        value_type = '차변'
    save_client_comparison_to_excel(df, account_list, value_type, base_dir=base_dir, output_filename=output_filename)

def run_menu_benford(df, base_dir=None, output_filename=None):
    print("\n   [벤포드 분석] 형식: 계정명:차변 또는 계정명:대변, 쉼표 구분")
    ben_in = input("   입력 (엔터=기본값): ").strip()
    ben_targets = []
    if ben_in:
        for x in ben_in.split(','):
            x = x.strip()
            if ':' in x:
                p = x.split(':', 1)
                d = p[1].strip() if len(p) > 1 else '차변'
                if d not in ('차변', '대변'):
                    d = '차변'
                ben_targets.append((p[0].strip(), d))
            elif x:
                ben_targets.append((x, '차변'))
    if not ben_targets:
        ben_targets = DEFAULT_BENFORD_TARGETS
    with _get_writer(base_dir, filename=output_filename) as w:
        ben_imgs = run_jet_analysis(df, w, ben_targets)
    path = _target_path(base_dir, filename=output_filename)
    if ben_imgs and os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        if '6_벤포드분석' in wb.sheetnames:
            ws = wb['6_벤포드분석']
            r = len(BENFORD_EXPLANATION) + 3
            for img_buf in ben_imgs:
                if img_buf:
                    ws.add_image(Image(img_buf), f'F{r}')
                    r += 25
        wb.save(path)
    fname = os.path.basename(path)
    print(f"   ✅ 벤포드 분석 결과가 {fname}에 저장되었습니다.")

def run_menu_monthly_full_account(df, base_dir=None, output_filename=None):
    """월별 전계정 분석 - 계정별·월별 차변/대변 합계"""
    with _get_writer(base_dir, filename=output_filename) as w:
        run_monthly_full_account(df, w)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 월별 전계정 분석이 {fname}에 저장되었습니다.")

def run_menu_benford_deviation_detail(df, base_dir=None, output_filename=None):
    """
    벤포드 분석에서 이탈이 큰 선행자릿수에 해당하는 거래 목록을 추출.
    별도 메뉴로 두어, 벤포드 분석 결과를 확인한 뒤 필요 시 상세 원인 파악용으로 실행.
    """
    print("\n   [벤포드 이탈 상세 추출] 벤포드 법칙에서 이탈이 큰 숫자에 해당하는 거래 목록을 추출합니다.")
    print("   형식: 계정명:차변 또는 계정명:대변 (쉼표로 복수 입력)")
    ben_in = input("   입력 (엔터=기본값 복리후생비:차변, 접대비:차변, 여비교통비:차변): ").strip()
    ben_targets = []
    if ben_in:
        for x in ben_in.split(','):
            x = x.strip()
            if ':' in x:
                p = x.split(':', 1)
                d = p[1].strip() if len(p) > 1 else '차변'
                if d not in ('차변', '대변'):
                    d = '차변'
                ben_targets.append((p[0].strip(), d))
            elif x:
                ben_targets.append((x, '차변'))
    if not ben_targets:
        ben_targets = DEFAULT_BENFORD_TARGETS

    thresh_in = input("   이탈 임계값 |차이| (엔터=0.03, 권장 0.02~0.05): ").strip()
    threshold = 0.03
    if thresh_in:
        try:
            t = float(thresh_in)
            if 0.001 <= t <= 0.3:
                threshold = t
        except ValueError:
            pass

    max_in = input("   숫자당 최대 건수 (엔터=500): ").strip()
    max_per_digit = 500
    if max_in:
        try:
            m = int(max_in)
            if 10 <= m <= 10000:
                max_per_digit = m
        except ValueError:
            pass

    benford_probs = {1: 0.301, 2: 0.176, 3: 0.125, 4: 0.097, 5: 0.079, 6: 0.067, 7: 0.058, 8: 0.051, 9: 0.046}
    all_detail = []
    summary_rows = []

    for acct, direction in ben_targets:
        target_col = COL_DEBIT if direction == '차변' else COL_CREDIT
        mask = _account_match_flexible(df[COL_ACCOUNT], acct)
        subset = df[mask & (df[target_col] > 0)].copy()
        n = len(subset)
        if n < BENFORD_MIN_ROWS:
            summary_rows.append({'계정': acct, '방향': direction, '선행자릿수': '-', '이탈정도': '-', '추출건수': '-', '전체건수': n, '비고': '데이터 부족'})
            continue

        subset['Digit'] = subset[target_col].apply(get_first_digit)
        digit_subset = subset[subset['Digit'] >= 1]['Digit']
        counts = digit_subset.value_counts(normalize=True).sort_index()
        deviant_digits = []
        for d in range(1, 10):
            actual = counts.get(d, 0.0)
            theory = benford_probs[d]
            diff = actual - theory
            if abs(diff) >= threshold:
                deviant_digits.append((d, diff, actual, theory))

        if not deviant_digits:
            summary_rows.append({'계정': acct, '방향': direction, '선행자릿수': '-', '이탈정도': '-', '추출건수': '-', '전체건수': n, '비고': '임계값 이상 이탈 없음'})
            continue

        deviant_set = {t[0] for t in deviant_digits}
        deviant_df = subset[subset['Digit'].isin(deviant_set)].copy()
        deviant_df['이탈정도'] = deviant_df['Digit'].map(lambda d: round(counts.get(d, 0) - benford_probs[d], 3))
        deviant_df = deviant_df.sort_values(by=[target_col], ascending=False)

        out_cols = [COL_DATE, COL_JOURNAL_ID, COL_ACCOUNT]
        if COL_CLIENT in deviant_df.columns:
            out_cols.append(COL_CLIENT)
        if COL_DESC in deviant_df.columns:
            out_cols.append(COL_DESC)
        out_cols.extend([target_col, 'Digit', '이탈정도'])

        for d in sorted(deviant_set):
            sub = deviant_df[deviant_df['Digit'] == d]
            sub = sub.head(max_per_digit)
            summary_rows.append({
                '계정': acct, '방향': direction, '선행자릿수': d,
                '이탈정도': round(counts.get(d, 0) - benford_probs[d], 3),
                '추출건수': len(sub), '전체건수': int((deviant_df['Digit'] == d).sum()), '비고': ''
            })
            export = sub[[c for c in out_cols if c in sub.columns]].copy()
            export['계정분류'] = f'{acct}({direction})'
            all_detail.append(export)

    if not all_detail and not summary_rows:
        print("   ⚠️ 추출할 데이터가 없습니다.")
        return

    path = _target_path(base_dir, filename=output_filename)
    with pd.ExcelWriter(path, engine='openpyxl') as w:
        info = [
            '[벤포드 이탈 상세 추출]',
            '이탈 임계값 이상인 선행자릿수(1~9)에 해당하는 거래를 금액순 상위 N건 추출한 결과입니다.',
            '원인 파악 시: 해당 숫자(예: 7)에 해당하는 금액(7만원대, 70만원대 등)의 거래처·적요·일자 패턴을 확인하세요.',
        ]
        pd.DataFrame({'내용': info}).to_excel(w, sheet_name='벤포드_이탈상세', startrow=0, index=False, header=False)
        if summary_rows:
            pd.DataFrame(summary_rows).to_excel(w, sheet_name='벤포드_이탈상세', startrow=len(info) + 2, index=False)
        if all_detail:
            combined = pd.concat(all_detail, axis=0, ignore_index=True)
            start = len(info) + 2 + (len(summary_rows) + 2 if summary_rows else 0)
            combined.to_excel(w, sheet_name='벤포드_이탈상세', startrow=start, index=False)

    fname = os.path.basename(path)
    print(f"   ✅ 벤포드 이탈 상세 추출 결과가 {fname}에 저장되었습니다.")

def run_menu_data_overview(df, base_dir=None, output_filename=None):
    with _get_writer(base_dir, filename=output_filename) as w:
        run_data_overview(df, w)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 데이터 개요가 {fname}에 저장되었습니다.")

def run_menu_account_list(df, base_dir=None, output_filename=None):
    with _get_writer(base_dir, filename=output_filename) as w:
        run_account_list(df, w)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 계정명 리스트가 {fname}에 저장되었습니다.")

def run_menu_employee_summary(df, base_dir=None, output_filename=None):
    with _get_writer(base_dir, filename=output_filename) as w:
        run_employee_summary(df, w)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 사원별 집계가 {fname}에 저장되었습니다.")

def run_menu_date_difference(df, base_dir=None, output_filename=None):
    print("\n   [일자차이 분석] 전표일자 vs 등록일자 차이 (일 단위)")
    s = input("   일수 입력 (엔터=건너뜀): ").strip()
    days = None
    if s:
        try:
            days = int(s)
            if days < 0:
                days = None
        except ValueError:
            pass
    with _get_writer(base_dir, filename=output_filename) as w:
        run_date_difference_analysis(df, w, days)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 일자차이 분석이 {fname}에 저장되었습니다.")

def run_menu_counterpart(df, base_dir=None, output_filename=None):
    print("\n   [상대계정 분석] 예: 접대비:차변")
    rel_in = input("   입력 (엔터=건너뜀): ").strip()
    acct, direction = None, '차변'
    if rel_in:
        if ':' in rel_in:
            p = rel_in.split(':', 1)
            acct = p[0].strip()
            direction = p[1].strip() if len(p) > 1 else '차변'
        else:
            acct = rel_in.strip()
    with _get_writer(base_dir, filename=output_filename) as w:
        run_counterpart_analysis(df, w, acct, direction)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 상대계정 분석이 {fname}에 저장되었습니다.")

def run_menu_keyword(df, base_dir=None, output_filename=None):
    key_in = input("   키워드 (쉼표 구분, 엔터=기본값): ").strip()
    keywords = [x.strip() for x in key_in.split(',') if x.strip()] if key_in else SEARCH_KEYWORDS
    with _get_writer(base_dir, filename=output_filename) as w:
        run_keyword_search(df, w, keywords)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 키워드 검색이 {fname}에 저장되었습니다.")

def run_menu_round_number(df, base_dir=None, output_filename=None):
    with _get_writer(base_dir, filename=output_filename) as w:
        run_round_number_analysis(df, w)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 라운드넘버 분석이 {fname}에 저장되었습니다.")

def run_menu_related_party(df, base_dir=None, output_filename=None):
    print("\n   [특수관계자 분석] 사용 중인 특수관계자 리스트:")
    for i, name in enumerate(RELATED_PARTIES, 1):
        print(f"      {i}. {name}")
    print(f"   (총 {len(RELATED_PARTIES)}개)")
    add_in = input("   추가할 특수관계자가 있으면 쉼표로 구분하여 입력하세요 (엔터=없음): ").strip()
    party_list = list(RELATED_PARTIES)
    if add_in:
        extras = [x.strip() for x in add_in.split(',') if x.strip()]
        for e in extras:
            if e and e not in party_list:
                party_list.append(e)
        if extras:
            print(f"   → 이번 분석에 추가: {', '.join(extras)} (총 {len(party_list)}개 대상)")
    print(f"   분석 대상 {len(party_list)}명으로 분석 실행합니다.")
    with _get_writer(base_dir, filename=output_filename) as w:
        run_related_party_analysis(df, w, parties=party_list)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 특수관계자 분석이 {fname}에 저장되었습니다.")

def run_menu_asset_liability(df, base_dir=None, output_filename=None):
    ast = input("   자산 계정 (쉼표 구분): ").strip().split(',')
    lbl = input("   부채 계정 (쉼표 구분): ").strip().split(',')
    assets = [x.strip() for x in ast if x.strip()]
    liabs = [x.strip() for x in lbl if x.strip()]
    with _get_writer(base_dir, filename=output_filename) as w:
        run_asset_liability_cross_check(df, w, assets, liabs)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 자산 vs 부채 교차가 {fname}에 저장되었습니다.")

def run_menu_revenue_expense(df, base_dir=None, output_filename=None):
    rev = input("   매출 계정 (쉼표 구분): ").strip().split(',')
    exp = input("   비용 계정 (쉼표 구분): ").strip().split(',')
    revs = [x.strip() for x in rev if x.strip()]
    exps = [x.strip() for x in exp if x.strip()]
    with _get_writer(base_dir, filename=output_filename) as w:
        run_revenue_expense_cross_check(df, w, revs, exps)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 매출 vs 비용 교차가 {fname}에 저장되었습니다.")

def run_menu_top_analysis(df, base_dir=None, output_filename=None, period_label=None):
    print("   계정:개수:차변/대변 (쉼표 구분, 방향 생략=둘다, 엔터=기본값)", flush=True)
    top_in = input("   입력: ").strip()
    config = []
    if top_in:
        for item in top_in.split(','):
            item = item.strip()
            if ':' in item:
                parts = item.split(':', 2)  # 계정, 개수, [차변|대변]
                acct = parts[0].strip()
                try:
                    n = int(parts[1].strip()) if len(parts) > 1 else 20
                except Exception:
                    n = 20
                dir_raw = parts[2].strip() if len(parts) > 2 else ''
                direction = 'both'
                if dir_raw == '차변':
                    direction = '차변'
                elif dir_raw == '대변':
                    direction = '대변'
                config.append((acct, n, direction))
            elif item:
                config.append((item, 20, 'both'))
    else:
        config = [(acct, n, 'both') for acct, n in USER_TOP_ANALYSIS_CONFIG]
    with _get_writer(base_dir, filename=output_filename) as w:
        run_user_defined_top_analysis(df, w, config, period_label=period_label)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 심층분석(계정별 Top)이 {fname}에 저장되었습니다.")

def run_menu_ai_preparation(df, base_dir=None, output_filename=None):
    """AI 계정별 분석: 거래처 익명화 + 계좌번호 마스킹 후 AI 전달용 파일·비밀_암호해독표 생성"""
    print("\n   [AI 계정별 분석] 거래처 익명화, 계좌번호 마스킹 후 AI 전달용 파일 생성")
    ai_in = input("   계정과목 (쉼표 구분, 예: 복리후생비, 접대비): ").strip()
    ai_targets = [x.strip() for x in ai_in.split(',') if x.strip()]
    if not ai_targets:
        print("     ⚠️ 계정을 입력하지 않아 건너뜁니다.")
        return
    base = base_dir or SCRIPT_DIR
    out_dir = os.path.normpath(os.path.join(base, OUT_DIR_NAME.strip()))
    os.makedirs(out_dir, exist_ok=True)
    ai_file = os.path.join(out_dir, "3_AI_업로드용_데이터.xlsx")
    key_file = os.path.join(out_dir, "4_비밀_암호해독표.xlsx")
    def _save_ai_results(writer, kf_path, data_df):
        for acct in ai_targets:
            run_ai_preparation(data_df, acct, writer)
        if GLOBAL_SAFE_MAP:
            pd.DataFrame(list(GLOBAL_SAFE_MAP.items()), columns=['실명', '가명']).to_excel(kf_path, index=False)

    try:
        with _get_writer(base_dir, filename="3_AI_업로드용_데이터.xlsx") as writer:
            _save_ai_results(writer, key_file, df)
        print(f"   ✅ AI 전달용 데이터: {os.path.basename(ai_file)}")
        if GLOBAL_SAFE_MAP:
            print(f"   ✅ 비밀_암호해독표: {os.path.basename(key_file)}")
    except PermissionError:
        alt_name = "3_AI_업로드용_데이터_백업.xlsx"
        alt_file = os.path.join(out_dir, alt_name)
        try:
            with _get_writer(base_dir, filename=alt_name) as writer:
                _save_ai_results(writer, os.path.join(out_dir, "4_비밀_암호해독표_백업.xlsx"), df)
            print(f"   ⚠️ 원본 파일이 사용 중이라 백업 파일로 저장했습니다.")
            print(f"   ✅ AI 전달용 데이터: {alt_name}")
        except Exception:
            print(f"   ⚠️ AI 분석 저장 실패: 해당 엑셀 파일이 다른 프로그램(예: Excel)에서 열려 있습니다.")
            print(f"      → '3_AI_업로드용_데이터.xlsx' 파일을 닫고 다시 실행하세요.")
    except Exception as e:
        print(f"   ⚠️ AI 분석 저장 실패: {e}")

def analyze_client_detail(df, account_names, client_names, value_type='both'):
    """
    특정 계정·거래처 조건에 맞는 전표 내역과 월별 합산 반환.
    - account_names, client_names: 리스트(여러 개 선택 시 하나라도 포함되면 해당).
    - value_type: '차변', '대변', 'both'(미선택=차변+대변 모두)
    반환: (detail_df, monthly_df)
    """
    if COL_ACCOUNT not in df.columns or COL_CLIENT not in df.columns or COL_DATE not in df.columns:
        return pd.DataFrame(), pd.DataFrame()
    if not client_names:
        return pd.DataFrame(), pd.DataFrame()
    # 계정: 비어 있으면 전체 계정(필터 없음), 있으면 하나라도 매칭되면 포함 (OR)
    if not account_names:
        mask_acct = pd.Series(True, index=df.index)
    else:
        mask_acct = pd.Series(False, index=df.index)
        for acct in account_names:
            if str(acct).strip():
                mask_acct = mask_acct | _account_match_flexible(df[COL_ACCOUNT], acct.strip())
    # 여러 거래처: 하나라도 포함되면 포함 (OR)
    client_col = df[COL_CLIENT].fillna('').astype(str)
    mask_client = pd.Series(False, index=df.index)
    for cli in client_names:
        if str(cli).strip():
            mask_client = mask_client | client_col.str.contains(cli.strip(), na=False, regex=False)
    filtered = df[mask_acct & mask_client].copy()
    if filtered.empty:
        return pd.DataFrame(), pd.DataFrame()
    if value_type == '차변':
        filtered = filtered[filtered[COL_DEBIT].fillna(0).astype(float) > 0]
    elif value_type == '대변':
        filtered = filtered[filtered[COL_CREDIT].fillna(0).astype(float) > 0]
    # value_type == 'both': 차변·대변 모두 포함 (추가 필터 없음)
    if filtered.empty:
        return pd.DataFrame(), pd.DataFrame()
    filtered['YM'] = pd.to_datetime(filtered[COL_DATE], errors='coerce').dt.strftime('%Y-%m')
    filtered = filtered[filtered['YM'].notna()]
    # 전표 내역용 컬럼: 구분·일자·전표번호·계정·차변·대변·거래처·적요 등 (있으면 사원명 포함)
    detail_cols = [c for c in [COL_DATE, COL_JOURNAL_ID, COL_ACCOUNT, COL_DEBIT, COL_CREDIT, COL_CLIENT, COL_DESC] if c in filtered.columns]
    if '구분' in filtered.columns:
        detail_cols.insert(0, '구분')
    if COL_EMPLOYEE in filtered.columns:
        detail_cols.append(COL_EMPLOYEE)
    detail_df = filtered[[c for c in detail_cols if c in filtered.columns]].copy()
    # 전표일자·전표번호 순 정렬 (관련 전표 전체 보기 쉽게)
    if COL_DATE in detail_df.columns:
        detail_df = detail_df.sort_values(by=[COL_DATE, COL_JOURNAL_ID] if COL_JOURNAL_ID in detail_df.columns else [COL_DATE], ascending=[True, True])
    # 월별 합산
    monthly = filtered.groupby('YM').agg({
        COL_DEBIT: 'sum',
        COL_CREDIT: 'sum',
    }).reset_index()
    monthly.columns = ['YM', '차변합계', '대변합계']
    monthly['합계'] = monthly['차변합계'] + monthly['대변합계']
    return detail_df, monthly

def run_menu_client_analysis(df, base_dir=None, output_filename=None):
    """거래처 분석: 계정·거래처(여러 개 선택 가능, 쉼표 구분) 전표 내역 + 월별 합산. 차변/대변 선택(미선택=둘 다)"""
    print("\n   [거래처 분석] 계정·거래처별 전표 내역 및 월별 합산")
    print("   계정·거래처는 쉼표로 구분하면 여러 개 선택됩니다. 계정 생략 시 해당 거래처의 전 계정 조회.")
    account_in = input("   계정과목 (쉼표 구분, 엔터=전체 계정): ").strip()
    account_list = [x.strip() for x in account_in.split(',') if x.strip()]
    if not account_list:
        print("   → 전체 계정으로 조회합니다.")
    client_in = input("   거래처명 (쉼표 구분, 부분 일치): ").strip()
    client_list = [x.strip() for x in client_in.split(',') if x.strip()]
    if not client_list:
        print("     ⚠️ 거래처명을 입력하지 않아 건너뜁니다.")
        return
    value_type = input("   금액 기준 (차변/대변/엔터=차변·대변 모두): ").strip()
    if value_type not in ('차변', '대변'):
        value_type = 'both'
    detail_df, monthly_df = analyze_client_detail(df, account_list, client_list, value_type)
    if detail_df.empty:
        print("     ⚠️ 해당 조건의 전표 내역이 없습니다.")
        return
    target_path = _target_path(base_dir, filename=output_filename)
    sheet_name = '거래처분석'
    label_type = '차변·대변 모두' if value_type == 'both' else value_type
    account_display = ', '.join(account_list) if account_list else '(전체 계정)'
    client_display = ', '.join(client_list)
    info_rows = pd.DataFrame([
        ['선택 조건', ''],
        ['계정과목', account_display],
        ['거래처명', client_display],
        ['금액 기준', label_type],
        ['전표 건수', str(len(detail_df)) + '건 (아래 전표 내역 참조)'],
    ])
    def _excel_val(val):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return ''
        if hasattr(val, 'item'):
            return val.item()
        if isinstance(val, pd.Timestamp):
            return str(val)
        return val

    # openpyxl로 직접 시트 기록 (전표 내역·월별 합산 모두 확실히 기록)
    if os.path.exists(target_path):
        wb = openpyxl.load_workbook(target_path)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name, 0)
    row = 1
    for r_row in dataframe_to_rows(info_rows, index=False, header=False):
        for c_idx, val in enumerate(r_row, 1):
            ws.cell(row=row, column=c_idx, value=_excel_val(val))
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="※ 전표 내역 (상세)")
    row += 2
    # 전표 내역: 헤더 + 데이터 행 (값은 _excel_val로 변환)
    for col_idx, col_name in enumerate(detail_df.columns, 1):
        ws.cell(row=row, column=col_idx, value=col_name)
    row += 1
    for _, r in detail_df.iterrows():
        for col_idx, val in enumerate(r, 1):
            ws.cell(row=row, column=col_idx, value=_excel_val(val))
        row += 1
    row += 1
    ws.cell(row=row, column=1, value="※ 월별 합산 금액")
    row += 2
    for r_row in dataframe_to_rows(monthly_df, index=False, header=True):
        for c_idx, val in enumerate(r_row, 1):
            ws.cell(row=row, column=c_idx, value=_excel_val(val))
        row += 1
    wb.save(target_path)
    fname = os.path.basename(target_path)
    print(f"   ✅ {fname} [거래처분석] 시트에 전표 내역 {len(detail_df)}건, 월별 합산 저장되었습니다.")

def run_menu_header_check(df, base_dir=None, output_filename=None):
    """data/current·data/previous 각 파일의 헤더 인식 결과를 출력 (전기 데이터 헤더 확인용)"""
    base_dir = base_dir or SCRIPT_DIR
    current_dir = os.path.join(base_dir, 'data', 'current')
    previous_dir = os.path.join(base_dir, 'data', 'previous')
    dtype_map = {COL_JOURNAL_ID: str}
    required = [(COL_DEBIT, '차변'), (COL_CREDIT, '대변'), (COL_ACCOUNT, '계정명'), (COL_CLIENT, '거래처명'), (COL_DATE, '전표일자')]

    print("\n   [데이터·헤더 확인] data/current(당기), data/previous(전기) 파일별 헤더 인식 결과")
    print("   " + "=" * 60)

    for label, dir_path in [('당기', current_dir), ('전기', previous_dir)]:
        print(f"\n   ▶ {label} 폴더: {dir_path}")
        if not os.path.isdir(dir_path):
            print(f"      ℹ️ 폴더 없음")
            continue
        files = sorted([f for f in os.listdir(dir_path) if (f.endswith('.xlsx') or f.endswith('.csv')) and not f.startswith('~$')])
        if not files:
            print(f"      ℹ️ 엑셀/CSV 파일 없음")
            continue
        for f in files:
            path = os.path.join(dir_path, f)
            ext = os.path.splitext(path)[1].lower()
            print(f"\n      📄 파일: {f}")
            try:
                if ext == '.xlsx':
                    trial, header_row = _read_excel_with_header_detection(path, dtype_map)
                    print(f"         헤더 행: {header_row + 1}번째 행 (엑셀 기준)")
                    print(f"         컬럼 수: {len(trial.columns)}")
                    cols = [str(c).strip() for c in trial.columns]
                    print(f"         컬럼 목록: {cols[:10]}{' ...' if len(cols) > 10 else ''}")
                    trial_norm = _normalize_debit_credit_columns(trial.copy())
                    ok = [name for col, name in required if col in trial_norm.columns]
                    miss = [name for col, name in required if col not in trial_norm.columns]
                    print(f"         필수 컬럼 인식: 차변/대변/계정명/거래처/전표일자 → OK: {ok}, 없음: {miss}")
                    if not trial_norm.empty:
                        first = trial_norm.iloc[0]
                        sample = {k: str(first.get(k, '-'))[:40] for k in [COL_ACCOUNT, COL_CLIENT, COL_DEBIT, COL_CREDIT] if k in trial_norm.columns}
                        print(f"         샘플 1행: {sample}")
                else:
                    # CSV: 본문 데이터 로드와 동일한 방식 (cp949 → utf-8, low_memory=False)
                    try:
                        trial = pd.read_csv(path, dtype=dtype_map, encoding='cp949', low_memory=False)
                    except UnicodeDecodeError:
                        trial = pd.read_csv(path, dtype=dtype_map, encoding='utf-8', low_memory=False)
                    except Exception as csv_err:
                        print(f"         ⚠️ CSV 읽기 실패: {csv_err}")
                        continue
                    if trial.empty:
                        trial = pd.DataFrame()
                    print(f"         (CSV) 헤더: 1번째 행")
                    print(f"         컬럼 수: {len(trial.columns)}")
                    print(f"         컬럼 목록: {list(trial.columns)[:10]}{' ...' if len(trial.columns) > 10 else ''}")
                    try:
                        trial_norm = _normalize_debit_credit_columns(trial.copy())
                    except Exception as norm_err:
                        print(f"         ⚠️ 컬럼 정규화 실패: {norm_err}")
                        trial_norm = trial
                    ok = [name for col, name in required if col in trial_norm.columns]
                    miss = [name for col, name in required if col not in trial_norm.columns]
                    print(f"         필수 컬럼 인식: OK: {ok}, 없음: {miss}")
            except Exception as e:
                print(f"         ⚠️ 읽기 실패: {type(e).__name__}: {e}")
    print("\n   " + "=" * 60)
    print("   ✅ 헤더 확인 완료. 위에서 전기 파일의 '헤더 행'과 '필수 컬럼 인식'을 확인하세요.\n")

# =============================================================================
# 📋 [계정별 거래처별 잔액증감분석] - 기초잔액(전기명세) + 당기증감 → 기말잔액
# =============================================================================

_PREV_BALANCE_SKIP_SHEETS = {'요약', 'FS v7.1', 'bs0930.', 'p930', 'BS1231', 'PI1231', 'Sheet1'}

def load_client_mapping(base_dir=None):
    """
    data/previous/거래처_매핑.csv 로드 → dict {분개장거래처: 전기명세거래처}.
    파일 없으면 빈 dict 반환.
    CSV 형식: 분개장거래처,전기명세거래처
    """
    base = base_dir or SCRIPT_DIR
    path = os.path.join(base, 'data', 'previous', '거래처_매핑.csv')
    if not os.path.exists(path):
        return {}, None
    df_map = None
    for enc in ['utf-8-sig', 'utf-8', 'cp949', 'euc-kr']:
        try:
            df_map = pd.read_csv(path, encoding=enc, dtype=str)
            break
        except Exception:
            continue
    if df_map is None or df_map.empty:
        return {}, path
    df_map = df_map.dropna(how='all')
    cols = [str(c).strip() for c in df_map.columns]
    df_map.columns = cols
    from_col = next((c for c in cols if '분개장' in c or '원본' in c), cols[0])
    to_col   = next((c for c in cols if '명세' in c or '매핑' in c or '전기' in c), cols[1] if len(cols) > 1 else cols[0])
    mapping = {
        str(k).strip(): str(v).strip()
        for k, v in zip(df_map[from_col], df_map[to_col])
        if pd.notna(k) and pd.notna(v) and str(k).strip() and str(v).strip()
    }
    return mapping, path

def load_previous_balance_excel(base_dir=None):
    """
    data/previous/ 폴더에서 '전기 계정별_거래처별명세.xlsx' 로드.
    반환: (DataFrame[계정명, 거래처명, 기초잔액], 파일경로). 파일 없으면 (None, None).
    """
    base = base_dir or SCRIPT_DIR
    prev_dir = os.path.join(base, 'data', 'previous')
    path = None
    if os.path.isdir(prev_dir):
        for f in sorted(os.listdir(prev_dir)):
            if f.startswith('~$') or not f.endswith('.xlsx'):
                continue
            if '계정별' in f or '명세' in f:
                path = os.path.join(prev_dir, f)
                break
        if path is None:
            for f in sorted(os.listdir(prev_dir)):
                if f.startswith('~$') or not f.endswith('.xlsx'):
                    continue
                path = os.path.join(prev_dir, f)
                break
    if not path or not os.path.exists(path):
        return None, None

    xl = pd.ExcelFile(path, engine='openpyxl')
    all_rows = []

    for sheet in xl.sheet_names:
        if sheet.strip() in _PREV_BALANCE_SKIP_SHEETS:
            continue
        try:
            raw = xl.parse(sheet, header=None, dtype=str)
            if raw.empty or len(raw) < 2:
                continue
            header = [str(v).strip() if pd.notna(v) else '' for v in raw.iloc[0]]
            if not any(h for h in header):
                continue
            data = raw.iloc[1:].reset_index(drop=True)
            acct_idx   = next((i for i, h in enumerate(header) if '계정' in h), 0)
            client_idx = next((i for i, h in enumerate(header) if '거래처' in h), 2)
            bal_idx    = next((i for i, h in enumerate(header) if '잔액' in h or '잔고' in h), 3)
            if bal_idx >= len(data.columns) or client_idx >= len(data.columns):
                continue
            acct_col   = data.iloc[:, acct_idx].astype(str).str.strip()
            client_col = data.iloc[:, client_idx].astype(str).str.strip()
            bal_series = _to_numeric_amount(data.iloc[:, bal_idx].astype(str))
            valid = (
                client_col.notna() &
                (client_col != 'nan') & (client_col != '') &
                (~client_col.str.contains(r'소계|합계|계\s*:', na=False, regex=True))
            )
            if not valid.any():
                continue
            sub = pd.DataFrame()
            sub['계정명']   = acct_col[valid].replace({'nan': pd.NA, '': pd.NA}).ffill().fillna(sheet).values
            sub['거래처명'] = client_col[valid].values
            sub['기초잔액'] = bal_series[valid].values
            sub = sub[sub['기초잔액'] != 0]
            if not sub.empty:
                all_rows.append(sub)
        except Exception:
            continue

    if not all_rows:
        return None, path

    result = pd.concat(all_rows, ignore_index=True)
    result['계정명']   = result['계정명'].astype(str).str.strip()
    result['거래처명'] = result['거래처명'].astype(str).str.strip()
    result['기초잔액'] = pd.to_numeric(result['기초잔액'], errors='coerce').fillna(0)
    result = result.groupby(['계정명', '거래처명'], as_index=False)['기초잔액'].sum()
    return result, path


def run_menu_balance_movement_analysis(df, base_dir=None, output_filename=None):
    """
    계정별 거래처별 잔액증감분석.
    기초잔액(전기 계정별_거래처별명세.xlsx) + 당기 분개장 증감 → 기말잔액 출력.
    """
    print("\n   [계정별 거래처별 잔액증감분석]")
    print("   전기 계정별_거래처별명세.xlsx → 기초잔액  |  당기 분개장 → 당기증감  |  기말잔액 계산")

    print("\n   📂 전기 계정별_거래처별명세.xlsx 로드 중...", flush=True)
    prev_df, prev_path = load_previous_balance_excel(base_dir)
    if prev_df is None:
        print("   ⚠️ data/previous/ 폴더에서 계정별 명세 파일을 찾을 수 없습니다.")
        return
    n_accts = prev_df['계정명'].nunique()
    print(f"   ✅ 기초잔액 로드: {len(prev_df)}행, {n_accts}개 계정  ({os.path.basename(prev_path)})")

    # 계정 목록 표시 및 선택
    accts_available = sorted(prev_df['계정명'].unique().tolist())
    print(f"\n   ── 계정 목록 ({len(accts_available)}개) ──")
    for i, a in enumerate(accts_available, 1):
        total = prev_df[prev_df['계정명'] == a]['기초잔액'].sum()
        print(f"      {i:3}. {a}  ({total:>15,.0f})")

    acct_in = input("\n   분석 계정 번호/이름 (쉼표 구분, 엔터=전체): ").strip()
    if acct_in:
        selected = []
        for tok in acct_in.split(','):
            tok = tok.strip()
            if tok.isdigit():
                idx = int(tok) - 1
                if 0 <= idx < len(accts_available):
                    selected.append(accts_available[idx])
            else:
                matched = [a for a in accts_available if tok in a]
                selected.extend(matched)
        selected = list(dict.fromkeys(selected))
    else:
        selected = accts_available

    if not selected:
        print("   ⚠️ 선택된 계정이 없습니다.")
        return
    print(f"\n   선택: {', '.join(selected[:5])}{'  외 ' + str(len(selected)-5) + '개' if len(selected) > 5 else ''}  (총 {len(selected)}개)")

    # 계정 유형 선택
    print("\n   계정 유형 (기말잔액 계산 방식):")
    print("   1 = 자산     : 기말 = 기초 + 차변(증가) - 대변(감소)")
    print("   2 = 부채·자본 : 기말 = 기초 + 대변(증가) - 차변(감소)")
    print("   3 = 차변/대변 모두 표시 (기말잔액 = 기초 + 차변 - 대변)")
    type_in = input("   선택 (1/2/3, 엔터=1): ").strip() or '1'
    if type_in not in ('1', '2', '3'):
        type_in = '1'
    type_label = {'1': '자산', '2': '부채·자본', '3': '참고(차변-대변)'}[type_in]

    # 거래처 매핑 파일 로드
    mapping, mapping_path = load_client_mapping(base_dir)
    if mapping:
        print(f"   ✅ 거래처 매핑 적용: {len(mapping)}개 항목  ({os.path.basename(mapping_path)})")
    else:
        mapping_hint = os.path.join(base_dir or SCRIPT_DIR, 'data', 'previous', '거래처_매핑.csv')
        print(f"   ℹ️ 거래처 매핑 없음 (필요 시 {mapping_hint} 작성)")

    # 당기 분개장 데이터 (구분=당기 필터)
    if '구분' in df.columns:
        df_j = df[df['구분'].astype(str).str.strip() == '당기'].copy()
    else:
        df_j = df.copy()

    # 계정명 정규화 캐시 구축 (86만 행 반복 apply 대신 고유 계정명만 정규화 → 속도 최적화)
    if COL_ACCOUNT in df_j.columns:
        _unique_j_accts = df_j[COL_ACCOUNT].dropna().astype(str).str.strip().unique()
        _norm_j_cache = {a: _normalize_account_for_match(a) for a in _unique_j_accts}
        print(f"   ✅ 분개장 고유 계정명 캐시: {len(_norm_j_cache)}개")
    else:
        _norm_j_cache = {}

    # 결과 저장 준비
    target_path = _target_path(base_dir, filename=output_filename)
    if os.path.exists(target_path):
        wb = openpyxl.load_workbook(target_path)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    used_names = set(wb.sheetnames)

    saved_sheets = 0
    summary_rows = []
    all_unmatched = []  # 매핑 템플릿 생성용 (기초잔액만 있거나 증감만 있는 거래처)

    for acct in selected:
        # 1) 기초잔액 행 (해당 계정)
        prev_sub = prev_df[prev_df['계정명'] == acct][['거래처명', '기초잔액']].copy()

        # 2) 당기 분개장에서 해당 계정 필터 (캐시 기반 → isin 필터로 속도 최적화)
        if _norm_j_cache:
            norm_prev = _normalize_account_for_match(acct)
            matched_accts = [a for a, n in _norm_j_cache.items()
                             if n.startswith(norm_prev) or norm_prev in n or n in norm_prev]
            df_acct = df_j[df_j[COL_ACCOUNT].isin(matched_accts)].copy() if matched_accts else pd.DataFrame(columns=df_j.columns)
        else:
            df_acct = pd.DataFrame(columns=df_j.columns)

        if not df_acct.empty and COL_CLIENT in df_acct.columns:
            # 매핑 적용: 분개장 거래처명 → 전기명세 거래처명으로 대체
            client_mapped = df_acct[COL_CLIENT].map(mapping).fillna(df_acct[COL_CLIENT]) if mapping else df_acct[COL_CLIENT]
            df_acct = df_acct.assign(_client_key=client_mapped)
            jgrp = df_acct.groupby('_client_key', as_index=False).agg(
                당기차변=(COL_DEBIT, 'sum'),
                당기대변=(COL_CREDIT, 'sum'),
            )
            jgrp.columns = ['거래처명', '당기차변', '당기대변']
        else:
            jgrp = pd.DataFrame(columns=['거래처명', '당기차변', '당기대변'])

        # 3) 기초잔액 + 분개장 outer merge
        result = pd.merge(prev_sub, jgrp, on='거래처명', how='outer')
        result['기초잔액'] = result['기초잔액'].fillna(0)
        result['당기차변'] = result['당기차변'].fillna(0)
        result['당기대변'] = result['당기대변'].fillna(0)
        result['거래처명'] = result['거래처명'].fillna('(미기재)')

        # 미매핑 항목 수집: 기초잔액=0이면서 분개장 증감이 있는 행 (= 연결 안 된 분개장 거래처)
        unmatched = result[
            (result['기초잔액'] == 0) &
            ((result['당기차변'] != 0) | (result['당기대변'] != 0))
        ][['거래처명']].copy()
        if not unmatched.empty:
            unmatched.insert(0, '계정명', acct)
            all_unmatched.append(unmatched)

        # 4) 기말잔액 계산
        if type_in == '1':
            result['당기증가'] = result['당기차변']
            result['당기감소'] = result['당기대변']
        elif type_in == '2':
            result['당기증가'] = result['당기대변']
            result['당기감소'] = result['당기차변']
        else:
            result['당기증가'] = result['당기차변']
            result['당기감소'] = result['당기대변']
        result['기말잔액'] = result['기초잔액'] + result['당기증가'] - result['당기감소']

        out_cols = ['거래처명', '기초잔액', '당기증가', '당기감소', '기말잔액']
        if type_in == '3':
            result = result.rename(columns={'당기증가': '당기차변', '당기감소': '당기대변'})
            out_cols = ['거래처명', '기초잔액', '당기차변', '당기대변', '기말잔액']
        result = result[out_cols].sort_values(by='기말잔액', key=lambda s: s.abs(), ascending=False)

        # 합계행 추가
        total = {c: result[c].sum() if result[c].dtype != object else '' for c in out_cols}
        total['거래처명'] = '합  계'
        result = pd.concat([result, pd.DataFrame([total])], ignore_index=True)

        # 요약 집계 (합계 시트용)
        t = result.iloc[-1]
        summary_rows.append({
            '계정명': acct,
            '기초잔액': t['기초잔액'],
            out_cols[2]: t[out_cols[2]],
            out_cols[3]: t[out_cols[3]],
            '기말잔액': t['기말잔액'],
        })

        # 시트 이름
        base_sname = _safe_sheet_name(f"잔액_{acct}")
        sname = base_sname
        dup = 1
        while sname in used_names:
            sname = _safe_sheet_name(f"잔액_{acct}_{dup}", max_len=29)
            dup += 1
        used_names.add(sname)
        if sname in wb.sheetnames:
            wb.remove(wb[sname])
        ws = wb.create_sheet(sname)

        mapping_note = f"  |  거래처 매핑: {len(mapping)}개 적용" if mapping else "  |  거래처 매핑 없음"
        ws.cell(row=1, column=1, value=f"[{acct}]  거래처별 잔액증감분석  ({type_label})")
        ws.cell(row=2, column=1, value=f"기초잔액: 전기 명세  |  당기증감: 당기 분개장  |  기말잔액: 기초 ± 증감{mapping_note}")

        for c_idx, col_name in enumerate(result.columns, 1):
            ws.cell(row=4, column=c_idx, value=col_name)
        for r_idx, row_data in enumerate(result.itertuples(index=False), 5):
            for c_idx, val in enumerate(row_data, 1):
                if isinstance(val, float) and pd.isna(val):
                    ws.cell(row=r_idx, column=c_idx, value='')
                elif isinstance(val, (int, float)) and not isinstance(val, bool):
                    ws.cell(row=r_idx, column=c_idx, value=float(val))
                else:
                    ws.cell(row=r_idx, column=c_idx, value=str(val) if val else '')
        saved_sheets += 1

    # 요약 시트 (전체 계정 합계)
    if summary_rows:
        sname_sum = '잔액증감_요약'
        if sname_sum in wb.sheetnames:
            wb.remove(wb[sname_sum])
        ws_sum = wb.create_sheet(sname_sum, 0)
        ws_sum.cell(row=1, column=1, value=f"계정별 거래처 잔액증감분석 요약  ({type_label})")
        ws_sum.cell(row=2, column=1, value=f"기준: 전기 명세(기초) + 당기 분개장(증감)")
        sum_df = pd.DataFrame(summary_rows)
        for c_idx, col_name in enumerate(sum_df.columns, 1):
            ws_sum.cell(row=4, column=c_idx, value=col_name)
        for r_idx, row_data in enumerate(sum_df.itertuples(index=False), 5):
            for c_idx, val in enumerate(row_data, 1):
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    ws_sum.cell(row=r_idx, column=c_idx, value=float(val))
                else:
                    ws_sum.cell(row=r_idx, column=c_idx, value=str(val) if val else '')

    wb.save(target_path)
    fname = os.path.basename(target_path)
    print(f"   ✅ 잔액증감분석 완료: {fname}  ({saved_sheets}개 계정 시트 + 요약 시트, 유형: {type_label})")

    # 매핑 템플릿 CSV 생성 (미연결 분개장 거래처 목록)
    if all_unmatched:
        tmpl_df = pd.concat(all_unmatched, ignore_index=True).drop_duplicates(subset=['거래처명'])
        tmpl_df.insert(1, '분개장거래처', tmpl_df['거래처명'])
        tmpl_df.insert(2, '전기명세거래처', '')
        tmpl_df = tmpl_df[['계정명', '분개장거래처', '전기명세거래처']]
        tmpl_path = os.path.join(base_dir or SCRIPT_DIR, 'data', 'previous', '거래처_매핑_템플릿.csv')
        tmpl_df.to_csv(tmpl_path, index=False, encoding='utf-8-sig')
        n_unmatched = len(tmpl_df)
        print(f"   📝 매핑 템플릿 저장: {n_unmatched}개 미연결 거래처 → {tmpl_path}")
        print(f"      ✏️  '전기명세거래처' 열에 매핑 대상을 입력 후 '거래처_매핑.csv'로 저장하면 다음 실행 시 적용됩니다.")


def run_general_ledger(df, writer, account_name):
    """선택된 계정과목의 월별 차변합계/대변합계/차변전표건수/대변전표건수 추출
    연도가 2개 이상이면 연도별 비교 형식(월×연도)으로 출력"""
    print(f"   ▶ [총계정원장] '{account_name}' 월별 집계 중...")

    mask = _account_match_flexible(df[COL_ACCOUNT], account_name)
    filtered = df[mask].copy()

    if filtered.empty:
        print(f"     ⚠️ '{account_name}' 계정의 데이터가 없습니다.")
        sheet_name = _safe_sheet_name(f"총계정원장_{account_name}")
        pd.DataFrame({'안내': [f"'{account_name}' 계정의 데이터가 없습니다."]}).to_excel(
            writer, sheet_name=sheet_name, index=False
        )
        return

    filtered['YM'] = pd.to_datetime(filtered[COL_DATE], errors='coerce').dt.strftime('%Y-%m')
    filtered = filtered[filtered['YM'].notna()]
    filtered['Year'] = filtered['YM'].str[:4]
    filtered['Month'] = filtered['YM'].str[5:7].astype(int)

    years = sorted(filtered['Year'].unique())

    debit_rows  = filtered[filtered[COL_DEBIT]  > 0]
    credit_rows = filtered[filtered[COL_CREDIT] > 0]

    if len(years) >= 2:
        # ── 연도 비교 형식: 행=월(1~12), 열=연도별 차변/대변 ──
        debit_agg = debit_rows.groupby(['Year', 'Month']).agg(
            차변합계=(COL_DEBIT, 'sum'), 차변전표건수=(COL_DEBIT, 'count')).reset_index()
        credit_agg = credit_rows.groupby(['Year', 'Month']).agg(
            대변합계=(COL_CREDIT, 'sum'), 대변전표건수=(COL_CREDIT, 'count')).reset_index()
        monthly_ym = pd.merge(debit_agg, credit_agg, on=['Year', 'Month'], how='outer').fillna(0)

        result = pd.DataFrame({'Month': range(1, 13)})
        for year in years:
            yd = monthly_ym[monthly_ym['Year'] == year][
                ['Month', '차변합계', '대변합계', '차변전표건수', '대변전표건수']
            ].rename(columns={c: f'{year}_{c}' for c in ['차변합계', '대변합계', '차변전표건수', '대변전표건수']})
            result = pd.merge(result, yd, on='Month', how='left').fillna(0)

        for year in years:
            result[f'{year}_차변전표건수'] = result[f'{year}_차변전표건수'].astype(int)
            result[f'{year}_대변전표건수'] = result[f'{year}_대변전표건수'].astype(int)

        # 증감 컬럼 (연도 2개일 때만)
        if len(years) == 2:
            y0, y1 = years[0], years[1]
            result[f'증감_차변합계({y0}→{y1})'] = result[f'{y1}_차변합계'] - result[f'{y0}_차변합계']
            result[f'증감_대변합계({y0}→{y1})'] = result[f'{y1}_대변합계'] - result[f'{y0}_대변합계']

        result.insert(0, '월', result['Month'].apply(lambda x: f'{x}월'))
        result = result.drop(columns=['Month'])

        total = {'월': '합  계'}
        for col in result.columns[1:]:
            total[col] = result[col].sum()
        result = pd.concat([result, pd.DataFrame([total])], ignore_index=True)
        mode_label = f"연도 비교 ({' vs '.join(years)})"

    else:
        # ── 단일 연도: 기존 시계열 형식 ──
        debit_agg = debit_rows.groupby('YM').agg(
            차변합계=(COL_DEBIT, 'sum'), 차변전표건수=(COL_DEBIT, 'count')).reset_index()
        credit_agg = credit_rows.groupby('YM').agg(
            대변합계=(COL_CREDIT, 'sum'), 대변전표건수=(COL_CREDIT, 'count')).reset_index()
        result = pd.merge(debit_agg, credit_agg, on='YM', how='outer').fillna(0).sort_values('YM')
        result['차변전표건수'] = result['차변전표건수'].astype(int)
        result['대변전표건수'] = result['대변전표건수'].astype(int)
        result = result.rename(columns={'YM': '월'})[['월', '차변합계', '대변합계', '차변전표건수', '대변전표건수']]
        total = {'월': '합  계', '차변합계': result['차변합계'].sum(), '대변합계': result['대변합계'].sum(),
                 '차변전표건수': int(result['차변전표건수'].sum()), '대변전표건수': int(result['대변전표건수'].sum())}
        result = pd.concat([result, pd.DataFrame([total])], ignore_index=True)
        mode_label = '월별 시계열'

    matched_accounts = df[mask][COL_ACCOUNT].unique().tolist()
    matched_display  = ', '.join(matched_accounts[:5])
    if len(matched_accounts) > 5:
        matched_display += f' 외 {len(matched_accounts) - 5}개'

    safe_name  = re.sub(r'[\\/*?:\[\]]', '', account_name)[:15]
    sheet_name = _safe_sheet_name(f"총계정원장_{safe_name}")

    info_rows = [
        f"[총계정원장] {account_name}  ({mode_label})",
        f"매칭 계정: {matched_display}",
        f"총 전표 건수: {len(filtered)}건",
    ]
    pd.DataFrame({'내용': info_rows}).to_excel(
        writer, sheet_name=sheet_name, startrow=0, index=False, header=False
    )
    result.to_excel(writer, sheet_name=sheet_name, startrow=len(info_rows) + 1, index=False)

    print(f"     ✅ 총계정원장 완료: '{account_name}' → {mode_label}")


def run_menu_general_ledger(df, base_dir=None, output_filename=None):
    """총계정원장: 선택된 계정과목의 월별 차변/대변 합계 및 전표건수"""
    print("\n   [총계정원장] 계정과목의 월별 차변합계/대변합계/차변전표건수/대변전표건수를 추출합니다.")
    print("   쉼표로 여러 계정 입력 가능 (예: 보통예금, 외상매출금)")
    account_in = input("   계정과목 입력: ").strip()
    if not account_in:
        print("     ⚠️ 계정을 입력하지 않아 건너뜁니다.")
        return
    account_list = [x.strip() for x in account_in.split(',') if x.strip()]
    with _get_writer(base_dir, filename=output_filename) as w:
        for acct in account_list:
            run_general_ledger(df, w, acct)
    fname = os.path.basename(_target_path(base_dir, filename=output_filename))
    print(f"   ✅ 총계정원장이 {fname}에 저장되었습니다.")


def run_menu_all(df, base_dir=None, output_filename=None):
    """전체 분석 수행 - 각 분석별로 사용자가 계정/설정을 선택한 뒤 순차 실행 (전체는 JET_통합분석결과.xlsx)"""
    print("\n   [전체 분석 수행] 각 분석별로 설정을 입력한 뒤 순차 실행합니다.")
    all_keys = sorted([k for k in MENU.keys() if k not in ('1', '16', '17', '18', '20')], key=lambda x: int(x) if x.isdigit() else 0)  # 19 포함
    for key in all_keys:
        name, runner = MENU[key]
        print("\n" + "-" * 50)
        print(f"▶ [{name}]")
        runner(df, base_dir=base_dir, output_filename=None)
    print(f"\n   ✅ 전체 분석이 모두 완료되었습니다. 결과: {TARGET_EXCEL_NAME}")

# 메뉴 딕셔너리: 번호 -> (표시 이름, 실행 함수)
MENU = {
    '1': ('전체 분석 수행', run_menu_all),
    '2': ('거래처 전기/당기 비교', run_menu_client_comparison),
    '3': ('벤포드 분석', run_menu_benford),
    '4': ('데이터 개요', run_menu_data_overview),
    '5': ('계정명 리스트', run_menu_account_list),
    '6': ('사원별 집계', run_menu_employee_summary),
    '7': ('일자차이 분석', run_menu_date_difference),
    '8': ('상대계정 분석', run_menu_counterpart),
    '9': ('키워드 검색', run_menu_keyword),
    '10': ('라운드넘버 분석', run_menu_round_number),
    '11': ('특수관계자 분석', run_menu_related_party),
    '12': ('자산 vs 부채 교차', run_menu_asset_liability),
    '13': ('매출 vs 비용 교차', run_menu_revenue_expense),
    '14': ('심층분석 (계정별 Top)', run_menu_top_analysis),
    '15': ('AI 계정별 분석', run_menu_ai_preparation),
    '16': ('데이터·헤더 확인', run_menu_header_check),
    '17': ('거래처 분석', run_menu_client_analysis),
    '18': ('벤포드 이탈 상세 추출', run_menu_benford_deviation_detail),
    '19': ('월별 전계정 분석', run_menu_monthly_full_account),
    '20': ('잔액증감분석 (계정별 거래처별)', run_menu_balance_movement_analysis),
    '21': ('총계정원장', run_menu_general_ledger),
}

def main():
    print("\n" + "="*80)
    print("🚀 [JET] 대화형 분석 메뉴")
    print("="*80)
    print("📂 데이터 로드: data/current/ (당기), data/previous/ (전기)")
    print("   (분석 실행 시 data 폴더의 엑셀/CSV 파일은 닫아 두세요. 열려 있으면 로드 실패할 수 있습니다)")
    df = load_data()
    if df is None:
        print("❌ 로드된 데이터 없음")
        return
    print(f"   ✅ 통합 데이터: {len(df)}행 (구분 컬럼: 당기/전기)")
    df = _preprocess_df(df)
    base_dir = SCRIPT_DIR
    target_path = _target_path(base_dir)
    print(f"   📁 결과 저장: {target_path}")

    while True:
        print("\n" + "-"*80)
        print("=== 분석 메뉴 선택 ===")
        sorted_keys = sorted(MENU.keys(), key=lambda x: int(x) if x.isdigit() else 0)
        left_keys = [k for k in sorted_keys if k.isdigit() and 1 <= int(k) <= 10]
        right_keys = [k for k in sorted_keys if k.isdigit() and 11 <= int(k)]
        right_keys.append('0')
        def _menu_label(k):
            return MENU[k][0] if k in MENU else "분석 종료"
        def _display_width(s):
            return sum(2 if unicodedata.east_asian_width(c) in ('W', 'F') else 1 for c in str(s))
        left_display_w = 50
        for i in range(max(len(left_keys), len(right_keys))):
            left = f"  {left_keys[i]}. {_menu_label(left_keys[i])}" if i < len(left_keys) else ""
            right = f"  {right_keys[i]:>2}. {_menu_label(right_keys[i])}" if i < len(right_keys) else ""
            pad = left_display_w - _display_width(left)
            print(f"{left}{' '*max(0, pad)}{right}")
        print("-"*80)
        choice = input("실행할 번호를 입력하세요 (예: 2,3 또는 1=전체, 0=종료): ").strip().replace(' ', '')

        if not choice:
            print("   ℹ️ 입력이 없습니다. 다시 선택하세요. (종료하려면 0 입력)")
            continue
        if choice == '0':
            print("분석을 종료합니다.")
            break

        # 다중 선택 파싱: "2,3" -> ['2','3'], "02"->"2", 전각숫자→반각
        def _normalize_menu_key(s):
            s = str(s).strip()
            # 전각 숫자(０-９) → 반각
            fw = '０１２３４５６７８９'
            hw = '0123456789'
            for a, b in zip(fw, hw):
                s = s.replace(a, b)
            s = s.lstrip('0') or '0'  # "02" -> "2", "0" -> "0"
            return s

        raw_keys = [k.strip() for k in choice.split(',') if k.strip()]
        keys = [_normalize_menu_key(k) for k in raw_keys]
        keys = list(dict.fromkeys(keys))

        try:
            for key in keys:
                if key not in MENU:
                    print(f"   ⚠️ '{key}' 번호는 메뉴에 없습니다. 건너뜁니다.")
                    continue
                name, runner = MENU[key]
                print(f"\n▶ [{name}] 실행 중...", flush=True)
                # 기간 선택: 2(전기/당기 비교), 16(헤더 확인), 20(잔액증감분석) 제외한 모든 분석에서 먼저 표시
                if key not in ('1', '2', '16', '20'):
                    sys.stdout.flush()
                    df_to_use, period_label = _ask_period_filter(df)
                    sys.stdout.flush()
                    if period_label in ('당기만', '전기만'):
                        print(f"   ℹ️ 기간: {period_label}", flush=True)
                else:
                    df_to_use = df
                # 전체(1)가 아니면 해당 번호 전용 파일명 사용 (예: JET_통합분석_3_벤포드분석.xlsx)
                out_file = None if key == '1' else _menu_filename(key, name)
                if key == '14':
                    run_menu_top_analysis(df_to_use, base_dir=base_dir, output_filename=out_file, period_label=period_label)
                else:
                    runner(df_to_use, base_dir=base_dir, output_filename=out_file)
        except PermissionError:
            print("\n⛔ [오류] 엑셀 파일이 열려있습니다. 해당 엑셀 파일을 닫은 뒤 같은 메뉴 번호를 다시 선택하세요.")
            print("   (프로그램을 종료하지 않았으므로 데이터 재로드 없이 다시 시도할 수 있습니다.)")
            continue
        except Exception as e:
            print(f"\n⛔ [오류 발생] {e}")
            import traceback
            traceback.print_exc()
            # 오류 후에도 메뉴로 돌아가서 재시도 가능
            continue

        print("\n" + "="*80)
        print(f"🎉 선택한 분석이 완료되었습니다. 결과 파일: {target_path}")
        print("   다른 분석을 선택하세요. (종료하려면 0 입력)")
        print("="*80)

if __name__ == "__main__":
    main()